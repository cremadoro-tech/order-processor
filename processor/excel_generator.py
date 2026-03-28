"""Excel作業指示書生成エンジン

カテゴリ別にシートを作成し、レイアウト定義に従ってデータを配置。
100行ごとにPartを分割。Summaryシートで全シートの数量を集計。
"""

import io
import re
import math
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.settings_io import load_json
from processor.transfer_engine import (
    has_transfer_rules, apply_transfer_rules, get_headers_for_sheet
)
from processor.houjin3_splitter import split_houjin3
from processor.amazon_extractor import extract_amazon_attributes, _shorten_product_name, expand_multi_name_rows

LAYOUTS_FILE = "sheet_layouts.json"

# スタイル定義
HEADER_FONT = Font(name="游ゴシック", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="游ゴシック", size=10)
BARCODE_FONT = Font(name="游ゴシック", size=10)
FRONTIER_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
PINK_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _resolve_sheet_name(row):
    """元マクロと同じロジックで出力シート名を決定する。

    優先順位:
    1. 特殊分類（印影確認・販売課）が設定されていればそれを使う
    2. 製品カテゴリ → sheet_mapping.json で日本語シート名に変換
    3. マッチしなければ「未分類」
    """
    mapping = load_json("sheet_mapping.json")
    special = mapping.get("特殊分類（最優先）", {})
    hanko = mapping.get("印鑑系（product_db）", {})
    outsource = mapping.get("外注系（outsource_db）", {})
    default_name = mapping.get("デフォルト", "未分類")

    # 1. 特殊分類チェック
    special_class = str(row.get("特殊分類", "")).strip()
    if special_class and special_class in special:
        return special[special_class]

    # 2. 製品カテゴリからマッピング
    product_cat = str(row.get("製品カテゴリ", "")).strip()
    if product_cat:
        # 印鑑系
        if product_cat in hanko:
            return hanko[product_cat]
        # 外注系
        if product_cat in outsource:
            return outsource[product_cat]

    return default_name


def generate_workbook(df: pd.DataFrame) -> bytes:
    """処理済みDataFrameからExcel作業指示書を生成してバイト列で返す。

    元マクロと同じシート振り分け: 製品カテゴリ → 日本語シート名 × 単品/複数
    """
    settings = load_json(LAYOUTS_FILE)
    layouts = settings.get("layouts", {})
    max_rows = settings.get("max_rows_per_sheet", 100)
    barcode_pre = settings.get("barcode_prefix", "*")
    barcode_suf = settings.get("barcode_suffix", "*")

    wb = Workbook()
    wb.remove(wb.active)

    summary_data = []

    # 各行に出力シート名を付与
    df = df.copy()
    df["_出力シート"] = df.apply(_resolve_sheet_name, axis=1)

    # Amazon専用レイアウト
    amazon_layouts = settings.get("amazon_layouts", {})
    amazon_sheet_names = amazon_layouts.get("amazon_sheet_names", {})

    # シート名別に処理
    for sheet_base, group in df.groupby("_出力シート", sort=False):
        # Amazon/楽天混在の場合はソースで分離
        has_amazon = (group.get("ソース") == "amazon").any() if "ソース" in group.columns else False
        has_rakuten = (group.get("ソース") != "amazon").any() if "ソース" in group.columns else True

        # 楽天データ
        if has_rakuten:
            rakuten_group = group[group.get("ソース", "") != "amazon"] if has_amazon else group
            layout = layouts.get(sheet_base, layouts.get("_default", {}))
            if "単品複数" in rakuten_group.columns:
                for qty_type, sub_group in rakuten_group.groupby("単品複数", sort=False):
                    sub_group = sub_group.reset_index(drop=True)
                    sheet_name = f"{sheet_base}_{qty_type}"
                    _write_sheet(wb, sheet_name, sub_group, layout, max_rows, barcode_pre, barcode_suf, summary_data)
            else:
                rakuten_group = rakuten_group.reset_index(drop=True)
                _write_sheet(wb, sheet_base, rakuten_group, layout, max_rows, barcode_pre, barcode_suf, summary_data)

        # Amazonデータ（専用レイアウト・シート名 + 複数名行展開）
        if has_amazon:
            amazon_group = group[group.get("ソース") == "amazon"]

            # 複数名入れの行展開
            # 同一管理番号+同一出力シートの組み合わせでは1回だけ展開
            # （同一注文に複数SKUがあっても備考テキストは同じため）
            expanded_rows = []
            expanded_keys = set()
            for _, row in amazon_group.iterrows():
                goq = str(row.get("GoQ管理番号", ""))
                result_rows = expand_multi_name_rows(row)
                if len(result_rows) > 1:
                    key = (goq, sheet_base)
                    if key in expanded_keys:
                        continue
                    expanded_keys.add(key)
                expanded_rows.extend(result_rows)
            amazon_group = pd.DataFrame(expanded_rows).fillna("").reset_index(drop=True)

            amazon_sheet = amazon_sheet_names.get(sheet_base, sheet_base)
            amazon_layout = amazon_layouts.get(sheet_base, amazon_layouts.get("_default", {}))
            # $ref参照を解決（共通レイアウト定義の再利用）
            if "$ref" in amazon_layout:
                amazon_layout = amazon_layouts.get(amazon_layout["$ref"], amazon_layouts.get("_default", {}))

            # Amazonは配送区分でシート分割（単品=AmazonJP, 単品＋=フロンティア+, 複数=フロンティア）
            amazon_group["_amazon_split"] = amazon_group.apply(_amazon_split_key, axis=1)
            for split_key, sub_group in amazon_group.groupby("_amazon_split", sort=False):
                sub_group = sub_group.reset_index(drop=True)
                sheet_name = f"{amazon_sheet}{split_key}"
                _write_sheet(wb, sheet_name, sub_group, amazon_layout, max_rows, barcode_pre, barcode_suf, summary_data)

    # Summaryシートを先頭に追加
    _write_summary(wb, summary_data)

    # バイト列に変換
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_sheet(
    wb: Workbook,
    base_name: str,
    df: pd.DataFrame,
    layout: dict,
    max_rows: int,
    barcode_pre: str,
    barcode_suf: str,
    summary_data: list,
):
    """DataFrameをシートに書き込む。max_rows超過時はPart分割。"""
    # カテゴリ名を抽出（"ジョインティ_単品" → "ジョインティ"）
    category_name = base_name.rsplit("_", 1)[0] if "_" in base_name else base_name

    # 法人3本セット分割: 1行→3行に展開
    is_houjin3 = layout.get("houjin3_split", False)
    if is_houjin3:
        expanded_rows = []
        for _, row in df.iterrows():
            expanded_rows.extend(split_houjin3(row))
        if expanded_rows:
            df = pd.DataFrame(expanded_rows)
        else:
            return

    # 転送ルールが存在するカテゴリか判定
    # Amazon専用レイアウト（amazon_source付き）の場合は転送エンジンを使わない
    has_amazon_layout = any("amazon_source" in col for col in layout.get("columns", []))
    use_transfer_engine = has_transfer_rules(category_name) and not is_houjin3 and not has_amazon_layout

    if use_transfer_engine:
        # 転送エンジンのヘッダーからカラム定義を生成
        headers = get_headers_for_sheet(category_name)
        columns = [{"header": h, "source": h} for h in headers]
    else:
        columns = layout.get("columns", [])
        if not columns:
            columns = load_json(LAYOUTS_FILE).get("layouts", {}).get("_default", {}).get("columns", [])

    total_rows = len(df)
    if total_rows == 0:
        return

    num_parts = math.ceil(total_rows / max_rows)

    for part in range(num_parts):
        start = part * max_rows
        end = min(start + max_rows, total_rows)
        part_df = df.iloc[start:end].reset_index(drop=True)

        # シート名（31文字制限）
        if num_parts > 1:
            sheet_name = f"{base_name}_Part_{part + 1}"
        else:
            sheet_name = base_name
        sheet_name = sheet_name[:31]

        # 重複シート名の回避
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            for suffix in range(1, 100):
                candidate = f"{sheet_name[:28]}_{suffix}"
                if candidate not in existing:
                    sheet_name = candidate
                    break

        ws = wb.create_sheet(title=sheet_name)

        # ヘッダー行
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # データ行
        for row_idx, (_, data_row) in enumerate(part_df.iterrows(), 2):
            is_frontier = str(data_row.get("配送区分", "")) == "フロンティア行"
            _source = data_row.get("ソース", "")
            is_amazon = str(_source) == "amazon" if _source is not None and str(_source) != "nan" else has_amazon_layout

            if use_transfer_engine and not is_amazon:
                # 楽天用: 転送エンジンで全列の値を一括生成
                transfer_result = apply_transfer_rules(data_row, category_name, row_number=row_idx - 1)

            if is_amazon:
                # Amazon用: 専用抽出器で属性を取得
                # NaN対策: pandasのNaNをstr()すると"nan"になるため空文字に変換
                _opts = data_row.get("項目・選択肢", "")
                _opts = "" if pd.isna(_opts) else str(_opts)
                _pname = data_row.get("商品名", "")
                _pname = "" if pd.isna(_pname) else str(_pname)
                _orderer = data_row.get("注文者氏名", "")
                _orderer = "" if pd.isna(_orderer) else str(_orderer)
                amazon_attrs = extract_amazon_attributes(_opts, _pname, _orderer)
                product_cat = str(data_row.get("製品カテゴリ", "") or "")
                short_name = _shorten_product_name(_pname, product_cat)

            for col_idx, col_def in enumerate(columns, 1):
                header = col_def["header"]

                if is_amazon:
                    # Amazon専用の値マッピング
                    amazon_src = col_def.get("amazon_source", header)
                    value = _extract_amazon_value(
                        data_row, header, amazon_attrs, short_name,
                        row_idx - 1, barcode_pre, barcode_suf,
                        amazon_source=amazon_src,
                    )
                elif use_transfer_engine:
                    value = transfer_result.get(header, "")
                else:
                    value = _extract_value(data_row, col_def, row_idx - 1, barcode_pre, barcode_suf)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # フロンティア行は赤背景
                if is_frontier:
                    cell.fill = FRONTIER_FILL

                # 個数2以上はピンク
                if col_def["header"] == "個数":
                    try:
                        if int(value) >= 2:
                            cell.fill = PINK_FILL
                    except (ValueError, TypeError):
                        pass

        # 列幅の自動調整
        _auto_column_width(ws, len(columns))

        # 印刷設定
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Summary用データ
        qty_col = next((c for c in columns if c["header"] == "個数"), None)
        if qty_col:
            qty_total = 0
            for _, row in part_df.iterrows():
                try:
                    qty_total += int(row.get("個数", 1))
                except (ValueError, TypeError):
                    qty_total += 1
        else:
            qty_total = len(part_df)

        summary_data.append({
            "シート名": sheet_name,
            "件数": len(part_df),
            "数量合計": qty_total,
        })


def _amazon_split_key(row):
    """Amazon用のシート分割キーを決定。

    マクロ②のI列と同じ分類:
    1. ひとことメモに「複数」+「単品」→ 単品＋（フロンティア+）
    2. ひとことメモに「複数」→ 複数（フロンティア）
    3. 単品複数列が「複数」or「単品+」→ 複数 or 単品＋（同一注文に他商品）
    4. それ以外 → 単品（AmazonJP）
    """
    memo = str(row.get("ひとことメモ", ""))
    has_fukusu = "複数" in memo
    has_tanpin = "単品" in memo

    if has_fukusu and has_tanpin:
        return "単品＋"
    elif has_fukusu:
        return "複数"

    # ひとことメモに「複数」がなくても、単品複数列で判定
    qty_type = str(row.get("単品複数", ""))
    if qty_type == "複数":
        return "複数"
    if qty_type == "単品+":
        return "単品＋"

    return "単品"


def _extract_value(row: pd.Series, col_def: dict, row_number: int, barcode_pre: str, barcode_suf: str):
    """カラム定義に従ってデータを抽出"""
    source = col_def.get("source", "")
    extract = col_def.get("extract", "")

    # 連番
    if source == "_row_number":
        return row_number

    raw = str(row.get(source, "")) if source in row.index else ""

    # 抽出パターン
    if extract == "size_mm":
        # 商品名からサイズ部分のみ抽出
        match = re.search(r"(\d+\.?\d*)\s*(?:mm|ｍｍ|ミリ)", raw)
        return f"{match.group(1)}ｍｍ" if match else raw[:20]

    if extract == "product_short":
        # 商品名を短縮（【】内を除去、30文字まで）
        short = re.sub(r"【[^】]*】", "", raw)
        short = re.sub(r"\[.*?\]", "", short)
        short = re.sub(r"【?10％OFF.*?】?", "", short)
        return short.strip()[:30]

    if extract == "body_color":
        # ボディカラーを抽出
        match = re.search(r"(?:カラー|ボディカラー|ボディーカラー)[=：:]\s*(.+?)(?:\n|$)", raw)
        return match.group(1).strip() if match else ""

    if extract == "illustration":
        # イラストを抽出
        match = re.search(r"(?:イラスト|イラストスタンプ)[=：:]\s*(.+?)(?:\n|$)", raw)
        return match.group(1).strip() if match else ""

    if extract == "ink_color":
        # 墨色/インク色を抽出
        match = re.search(r"(?:墨色|薄墨|インク色)[=：:]\s*(.+?)(?:\n|$)", raw)
        return match.group(1).strip() if match else ""

    if extract == "sei":
        # 姓を抽出（スペース区切りの前半）
        parts = raw.split()
        return parts[0] if parts else raw

    if extract == "mei":
        # 名を抽出（スペース区切りの後半）
        parts = raw.split()
        return parts[1] if len(parts) >= 2 else ""

    # バーコード列はプレフィックス/サフィックス付与
    if col_def["header"] == "バーコード" and raw:
        return f"{barcode_pre}{raw}{barcode_suf}"

    return raw


def _extract_amazon_value(row, header, amazon_attrs, short_name, row_number, barcode_pre, barcode_suf, amazon_source=""):
    """Amazon用の値取得。amazon_sourceキーで何のデータを取るか指定。"""
    src = amazon_source or header

    # 空文字指定 = 空欄出力（Amazon備考欄は作成内容抽出に使ったので空にする）
    if src == "":
        return ""
    if src in ("備考", "備考欄"):
        return ""

    # 連番（永江用Amazon等）
    if src == "_row_number":
        return row_number

    # 固定値パターン（_fixed:XXX）
    if src.startswith("_fixed:"):
        return src[7:]

    # 短縮商品名
    if src == "short_name":
        return short_name

    # 行展開された場合の優先値
    if src == "作成名" and "_expanded_name" in row.index and str(row.get("_expanded_name", "")):
        return str(row.get("_expanded_name", ""))
    if src == "書体" and "_expanded_font" in row.index and str(row.get("_expanded_font", "")):
        return str(row.get("_expanded_font", ""))

    # GoQ管理番号バーコード（*FA2-742* 形式）
    if src == "GoQ管理番号_barcode":
        val = row.get("GoQ管理番号", "")
        goq = "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val)
        return f"{barcode_pre}{goq}{barcode_suf}" if goq else ""

    # ゴム印専用: 内容（住所テキスト。属性行を除去）
    if src == "gomu_content":
        val = row.get("項目・選択肢", "")
        # 項目・選択肢が空の場合は備考にフォールバック
        if pd.isna(val) or not str(val).strip():
            val = row.get("備考", "")
        if pd.isna(val) or not str(val).strip():
            return ""
        orderer = row.get("注文者氏名", "")
        orderer = "" if pd.isna(orderer) else str(orderer).strip()
        return _clean_gomu_content(str(val), orderer)

    # ゴム印専用: サイズ（完成データ形式: "最大4行(60mm×20mm" or "62mm×15mm" or "20mm×60mm"）
    if src == "gomu_size":
        pname = row.get("商品名", "")
        pname = "" if pd.isna(pname) else str(pname)
        opts = row.get("項目・選択肢", "")
        opts = "" if pd.isna(opts) else str(opts)

        # 備考先頭のサイズを最優先（「60mmx25mm」等、お客様が選択したサイズ）
        m = re.match(r"\s*(\d+mm\s*[×xX]\s*\d+mm)", opts)
        if m:
            return m.group(1).strip()
        # 商品名の括弧から「最大N行(NNmm×NNmm)」を取得（Amazon選べるサイズ商品）
        m = re.search(r"\(?(最大\d行\(\d+mm[×x]\d+mm)\)?", pname)
        if m:
            return m.group(1)
        # 備考から「最大N行(NNmm×NNmm」
        m = re.search(r"(最大\d行\(\d+mm[×x]\d+mm)", opts)
        if m:
            return m.group(1)
        # 備考から「サイズ=NNmm×NNmm」or「印面サイズ:NNmm×NNmm」
        m = re.search(r"(?:サイズ|印面サイズ)[=:：]\s*(.+?)(?:\n|$)", opts)
        if m:
            return m.group(1).strip()
        # 商品名から NNmm×NNmm
        m = re.search(r"(\d+mm[×x]\d+mm)", pname)
        if m:
            return m.group(1)
        # フォールバック: amazon_attrsのサイズ
        return amazon_attrs.get("サイズ", "")

    # ゴム印専用: 配置（ヨコ/タテ）
    if src == "gomu_direction":
        opts = row.get("項目・選択肢", "")
        opts = "" if pd.isna(opts) else str(opts)
        pname = row.get("商品名", "")
        pname = "" if pd.isna(pname) else str(pname)
        return _extract_gomu_direction(opts, pname)

    # おなまえスタンプ専用: ひらがな/漢字/ローマ字/イラスト抽出
    if src.startswith("onamae_"):
        opts = row.get("項目・選択肢", "")
        opts = "" if pd.isna(opts) else str(opts)
        pname = row.get("商品名", "")
        pname = "" if pd.isna(pname) else str(pname)
        return _extract_onamae_field(src, opts, pname, row)

    # Amazon抽出結果
    if src in ("書体", "カラー", "作成名", "サイズ", "配置"):
        return amazon_attrs.get(src, "")

    def _safe_str(val):
        return "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val)

    # 元データから直接取得
    if src == "個数":
        val = _safe_str(row.get("個数", "1")) or "1"
        return val.replace(".0", "") if val.endswith(".0") else val
    if src == "GoQ管理番号":
        return _safe_str(row.get("GoQ管理番号", ""))
    if src == "ひとことメモ":
        return _safe_str(row.get("ひとことメモ", ""))
    if src == "単品複数":
        # AmazonJP / フロンティア / フロンティア+ の形式（マクロ③と一致）
        memo = _safe_str(row.get("ひとことメモ", ""))
        has_fukusu = "複数" in memo
        has_tanpin = "単品" in memo
        if has_fukusu and has_tanpin:
            return "フロンティア+"
        elif has_fukusu:
            return "Amazonフロンティア"
        qty = _safe_str(row.get("単品複数", ""))
        if qty == "複数":
            return "Amazonフロンティア"
        if qty == "単品+":
            return "フロンティア+"
        return "AmazonJP"
    if src == "注文者氏名":
        return str(row.get("注文者氏名", ""))
    if src == "セット":
        name = str(row.get("商品名", ""))
        if "2点セット" in name:
            return "2点セット"
        if "3点セット" in name:
            return "3点セット"
        return ""

    # フォールバック
    if src in row.index:
        val = row.get(src, "")
        return "" if pd.isna(val) else str(val)

    return ""


def _clean_gomu_content(text, orderer_name=""):
    """ゴム印の内容から属性行（書体・向き・サイズ等）を除去し、住所テキストのみ返す。

    入力例: "名入れ書体：【明朝体】 文字の向き：【ヨコ】 名入れ文字： 【〒592-8345 堺市西区浜寺昭和町1丁150-3 山本晃也】 山本晃也"
    出力: "〒592-8345 堺市西区浜寺昭和町1丁150-3 山本晃也"
    """
    # 【名入れ文字：】内のテキストを優先抽出
    m = re.search(r"名入れ文字[：:]\s*【([^】]+)】", text)
    if m:
        return m.group(1).strip()

    # 備考フォールバック時: タイムスタンプデータの除去
    text = re.sub(r"発送予定日\s+出荷予定日.*?23:59:59\s*", "", text, flags=re.DOTALL).strip()
    text = re.sub(r"お届け予定日.*?23:59:59\s*", "", text, flags=re.DOTALL).strip()
    text = re.sub(r"ギフトメッセージ\s*", "", text).strip()

    # 先頭の属性テキストを一括除去（「明朝体 横書き 60×25 〒...」パターン）
    _font_names = r"(?:楷書体|明朝体|行書体|隷書体|古印体|てん書体|印相体|角ゴシック体|丸ゴシック体|ゴシック体|クラフト体)"
    # 先頭の「・書体名[/スペース]・方向 サイズ」を除去（複数パターン対応）
    text = re.sub(
        rf"^[・\s]*{_font_names}[/\s　]*(?:[・]*(?:ヨコ|タテ|横書き|縦書き|縦判|ヨコ向き|タテ向き|横向き|縦向き))?[/\s　]*(?:\d+[×x]\d+)?[/\s　]*",
        "", text
    ).strip()
    # 「・横向き」「・縦向き」のみが先頭に残る場合
    text = re.sub(r"^[・\s]*(?:横向き|縦向き|ヨコ向き|タテ向き)\s*", "", text).strip()

    # 行ごとに処理して属性行を除去
    lines = text.split("\n")
    result = []
    remove_patterns = [
        r"^名入れ書体[：:]",
        r"^書体[：:]",
        r"^字体[：:]",
        r"^文字の向き[：:]",
        r"^文字の配置[：:]",
        r"^(?:向き|配置)[：:]",
        r"^印面に入れる内容[：:]$",
        r"^サイズ[=：:]",
        r"^印面サイズ[=：:]",
        r"^最大\d行\(",
        r"^\d+～\d+営業日",
        r"^[-—─]{3,}",
        r"^備考[=：:]",
        r"^ヨコ向き$",
        r"^タテ向き$",
        r"^横書き$",
        r"^縦書き$",
        r"^縦判$",
        r"^内容[：:]\s*$",
    ]
    font_only_pat = re.compile(
        rf"^[・]*{_font_names}[/]?$"
    )
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        skip = False
        for pat in remove_patterns:
            if re.search(pat, stripped):
                skip = True
                break
        if font_only_pat.match(stripped):
            skip = True
        if "(addr)" in stripped:
            skip = True
        if not skip:
            # 行内の属性プレフィックスを除去
            cleaned = re.sub(r"^印面に入れる内容[：:]\s*", "", stripped)
            cleaned = re.sub(r"^彫刻内容[=：:]\s*", "", cleaned)
            cleaned = re.sub(r"^内容[：:]\s*", "", cleaned)
            # インラインの「書体：XX」「向き：XX」「サイズ XX」を除去
            cleaned = re.sub(r"書体\s*[：:]\s*\S+\s*", "", cleaned)
            cleaned = re.sub(r"向き\s*[：:]\s*\S+\s*", "", cleaned)
            cleaned = re.sub(r"サイズ\s+\d+mm[×x]\d+mm\s*", "", cleaned)
            # 「彫刻名\s*：XX」パターンを除去
            cleaned = re.sub(r"彫刻名\s*[：:]\s*\S+\s*", "", cleaned)
            cleaned = cleaned.strip()
            if cleaned:
                result.append(cleaned)

    content = "\n".join(result)

    # 末尾の注文者名を除去（住所テキストの後にAmazon注文者名が追加されるパターン）
    if orderer_name and content:
        content = _remove_trailing_orderer(content, orderer_name)

    return content


def _remove_trailing_orderer(content, orderer_name):
    """内容末尾の注文者名を除去する。

    Amazonゴム印の備考パターン:
    「〒XXX-XXXX 住所 会社名 TEL XXX 注文者名」
    → 末尾の注文者名を除去
    """
    if not orderer_name or not content:
        return content

    # 注文者名のバリエーションを生成
    orderer_clean = orderer_name.strip()
    orderer_nospace = re.sub(r"[\s　]+", "", orderer_clean)
    orderer_parts = re.split(r"[\s　]+", orderer_clean)
    orderer_sei = orderer_parts[0] if orderer_parts else ""

    # 末尾一致チェック用の候補リスト（長い順）
    candidates = []
    if orderer_clean:
        candidates.append(orderer_clean)         # "山下 峰"
    if orderer_nospace and orderer_nospace != orderer_clean:
        candidates.append(orderer_nospace)         # "山下峰"
    # 姓のみは短すぎるので単独では使わない（誤除去防止）

    for name in candidates:
        # エスケープしてパターン生成
        escaped = re.escape(name)
        # スペースバリエーション（「山下 峰」も「山下　峰」もマッチ）
        flexible = re.sub(r"\\[\s　]+", r"[\\s　]*", escaped)
        # 末尾にマッチ（前にスペースがある）
        pattern = rf"[\s　]+{flexible}\s*$"
        new_content = re.sub(pattern, "", content)
        if new_content != content:
            return new_content.strip()

    # 別パターン: 同じ名前が重複（"森兼一郎 森兼一郎"）
    content_flat = re.sub(r"\s+", " ", content).strip()
    for name in candidates:
        escaped = re.escape(name)
        flexible = re.sub(r"\\[\s　]+", r"[\\s　]*", escaped)
        if re.search(rf"{flexible}\s+{flexible}\s*$", content_flat):
            # 末尾の重複を1つ除去
            new_content = re.sub(rf"\s+{flexible}\s*$", "", content)
            if new_content != content:
                return new_content.strip()

    return content


def _extract_gomu_direction(text, product_name):
    """ゴム印の配置（ヨコ/タテ）をAmazon備考から抽出。

    完成データの配置列: "ヨコ" or "タテ"
    """
    if not text:
        return ""

    # パターン1: 文字の向き：【ヨコ】 or 文字の配置：【ヨコ書き】
    m = re.search(r"(?:文字の向き|文字の配置|向き|配置)[：:=]\s*【?([^】\n]+)】?", text)
    if m:
        d = m.group(1).strip()
        if "ヨコ" in d or "横" in d:
            return "ヨコ"
        if "タテ" in d or "縦" in d:
            return "タテ"
        return d

    # パターン2: 単独の【ヨコ】【タテ】
    if re.search(r"【ヨコ】|【横】", text):
        return "ヨコ"
    if re.search(r"【タテ】|【縦】", text):
        return "タテ"

    # パターン3: 向き：ヨコ / 向き：タテ
    m = re.search(r"向き[：:]\s*(ヨコ|横|タテ|縦)", text)
    if m:
        d = m.group(1)
        return "ヨコ" if d in ("ヨコ", "横") else "タテ"

    # パターン4: 商品名にヨコ/タテの指定
    if "ヨコ" in product_name or "横" in product_name:
        return "ヨコ"
    if "タテ" in product_name or "縦" in product_name:
        return "タテ"

    # パターン5: テキスト内に直接書かれている
    if re.search(r"(?<![角丸])ヨコ|横書き", text):
        return "ヨコ"
    if re.search(r"タテ|縦書き|縦判", text):
        return "タテ"

    return ""


def _extract_onamae_field(field, opts, product_name, row=None):
    """おなまえスタンプ用フィールド抽出。

    項目・選択肢パターン:
    パターンA: 名入れ文字（ひらがな）：【むらかみ りゅうせい】名入れ文字（漢字）：【村上 龍誠】名入れ文字（ローマ字）：【MURAKAMI RYUSEI】
    パターンB: 作成するお名前】：みずの ゆめの（ひらがなのみ）
    イラスト: 028.やきゅう / イラスト：【008.みつばち】 / 【イラスト】：007
    """
    if field == "onamae_product":
        # 商品名を短縮（"14点セット", "8点セット" 等）
        # 9点セット = 8点スタンプ+ボックス付き → 製造上は8点セット
        m = re.search(r"(\d+)点セット", product_name)
        if m:
            n = int(m.group(1))
            if n == 9:
                n = 8  # 9点=8点+ボックス
            elif n == 15:
                n = 14  # 15点=14点+ボックス
            return f"{n}点セット"
        if "入園" in product_name:
            return "入園セット"
        return product_name[:20]

    if field == "onamae_split":
        # 東京製版用: 「単品」「フロンティア」ラベル
        if row is None:
            return "単品"
        memo = str(row.get("ひとことメモ", ""))
        memo = "" if memo == "nan" else memo
        has_fukusu = "複数" in memo
        has_tanpin = "単品" in memo
        if has_fukusu and has_tanpin:
            return "フロンティア+"
        elif has_fukusu:
            return "フロンティア"
        qty = str(row.get("単品複数", ""))
        qty = "" if qty == "nan" else qty
        if qty == "複数":
            return "フロンティア"
        if qty == "単品+":
            return "フロンティア+"
        return "単品"

    if field == "onamae_size":
        # 選べるサイズ（基本的に空、DXセット等で値あり）
        m = re.search(r"選べるサイズ[=：:]\s*(.+?)(?:\n|$)", opts)
        return m.group(1).strip() if m else ""

    if field == "onamae_hiragana":
        return _extract_onamae_name(opts, "hiragana")

    if field == "onamae_kanji":
        return _extract_onamae_name(opts, "kanji")

    if field == "onamae_roman":
        return _extract_onamae_name(opts, "roman")

    if field == "onamae_illust":
        return _extract_onamae_illust(opts)

    return ""


# イラスト条件リストキャッシュ
_illust_mapping_cache = None

def _get_illust_mapping():
    """イラスト条件リストを読み込み"""
    global _illust_mapping_cache
    if _illust_mapping_cache is None:
        try:
            data = load_json("onamae_illust_mapping.json")
            _illust_mapping_cache = data.get("mapping", {})
        except Exception:
            _illust_mapping_cache = {}
    return _illust_mapping_cache


def _lookup_illust(keyword):
    """キーワードから条件リストでイラスト名を検索"""
    mapping = _get_illust_mapping()
    if not keyword:
        return ""
    kw = keyword.strip()
    # 完全一致
    if kw in mapping:
        return mapping[kw]
    # 全角数字→半角変換して再検索
    zen = "０１２３４５６７８９"
    han = "0123456789"
    for z, h in zip(zen, han):
        kw = kw.replace(z, h)
    if kw in mapping:
        return mapping[kw]
    # 部分一致（キーワードがテキストに含まれるか）
    for mk, mv in mapping.items():
        if mk in kw or kw in mk:
            return mv
    return ""


def _extract_onamae_name(opts, name_type):
    """おなまえスタンプ用 ひらがな/漢字/ローマ字を抽出。

    name_type: "hiragana", "kanji", "roman"

    対応パターン:
    A: 名入れ文字（ひらがな）：【XX】名入れ文字（漢字）：【YY】名入れ文字（ローマ字）：【ZZ】
    B: 【ひらがな：XX】【漢字：YY】【ローマ字(大文字のみ)：ZZ】
    C: 名入れ文字：【XX】（ひらがなのみ、8点セット等）
    D: 作成するお名前】：XX
    E: 016、漢字名、ひらがな名、ローマ字名（カンマ区切り+番号プレフィックス）
    F: 【018.くま】【えぐちせしる】（括弧のみ）
    G: いちご/015. 名前 ひらがな名（スラッシュ+名前キーワード）
    H: フリーテキスト（先頭ひらがな名 注文者漢字名）
    """
    if not opts:
        return ""

    # === パターンA: 名入れ文字（ひらがな）：【XX】 ===
    label_map = {
        "hiragana": r"名入れ文字[（(]ひらがな[）)]",
        "kanji": r"名入れ文字[（(]漢字[）)]",
        "roman": r"名入れ文字[（(]ローマ字[）)]",
    }
    m = re.search(label_map[name_type] + r"[：:＝=]*\s*【([^】]+)】", opts)
    if m:
        return _clean_onamae_val(m.group(1), name_type)

    # === パターンB: 【ひらがな：XX】 ===
    label_map_b = {
        "hiragana": r"ひらがな",
        "kanji": r"漢字",
        "roman": r"ローマ字(?:\(大文字のみ\))?",
    }
    m = re.search(r"【" + label_map_b[name_type] + r"[：:]\s*([^】]+)】", opts)
    if m:
        return _clean_onamae_val(m.group(1), name_type)

    # === パターンE: 016、漢字名、ひらがな名、ローマ字名 ===
    m = re.match(r"\d{2,3}[、,]\s*(.+)", opts.split("\n")[0])
    if m:
        parts = re.split(r"[、,]\s*", m.group(1))
        # parts: [漢字, ひらがな, ローマ字, 注文者名] or [ひらがな, 漢字, ローマ字, ...]
        # 文字種で判定
        if len(parts) >= 3:
            classified = _classify_name_parts(parts[:3])
            if name_type in classified:
                return _clean_onamae_val(classified[name_type], name_type)

    # ひらがな専用の追加パターン
    if name_type == "hiragana":
        # パターンC: 名入れ文字：【XX】（ひらがな指定なし）
        m = re.search(r"名入れ文字[：:]\s*【([^】]+)】", opts)
        if m:
            return _clean_onamae_val(m.group(1), name_type)
        # パターンD: 作成するお名前】：XX
        m = re.search(r"作成する(?:お名前|おなまえ)[】]?[：:]\s*(.+?)(?:\n|【|$)", opts)
        if m:
            return _clean_onamae_val(m.group(1), name_type)
        # パターンF: 【018.くま】【えぐちせしる】→ 2番目の括弧内
        m = re.match(r"【[^】]*】\s*【([^】]+)】", opts)
        if m:
            val = m.group(1).strip()
            if re.search(r"[\u3040-\u309F]", val):
                return val
        # パターンG: いちご/015. 名前 ひらがな名
        m = re.search(r"名前[\s　]+([^\n]+)", opts)
        if m:
            val = m.group(1).strip()
            # ひらがな部分だけ取得
            m2 = re.match(r"([\u3040-\u309F\s　]+)", val)
            if m2:
                return m2.group(1).strip()
        # パターンH: フリーテキスト（先頭ひらがな名）
        first_line = opts.split("\n")[0].strip()
        # イラスト番号/名を除去
        first_line = re.sub(r"^【[^】]*】\s*", "", first_line)
        first_line = re.sub(r"^\d{2,3}[.\s]+\S+\s*", "", first_line).strip()
        first_line = re.sub(r"^\S+/\d{2,3}[.\s]*", "", first_line).strip()
        if first_line:
            m = re.match(r"([\u3040-\u309F\s　]+)", first_line)
            if m:
                name = m.group(1).strip()
                name = re.sub(r"[\s　]+(でお|で|お|に|を)\s*$", "", name).strip()
                if len(name) >= 2:
                    return name

    return ""


def _classify_name_parts(parts):
    """名前パーツを文字種で分類（ひらがな/漢字/ローマ字）"""
    result = {}
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if re.match(r"^[A-Za-zＡ-Ｚ\s　]+$", p):
            result["roman"] = p
        elif re.search(r"[\u3040-\u309F]", p) and not re.search(r"[\u4E00-\u9FFF]", p):
            result["hiragana"] = p
        elif re.search(r"[\u4E00-\u9FFF]", p):
            result["kanji"] = p
    return result


def _clean_onamae_val(val, name_type):
    """おなまえスタンプの値をクリーニング"""
    val = val.strip()
    if name_type == "roman":
        # 全角英字→半角
        val = val.translate(str.maketrans(
            'ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ',
            'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        ))
        # 注意書き除去
        val = re.sub(r"※.*$", "", val).strip()
    return val


def _extract_onamae_illust(opts):
    """おなまえスタンプのイラスト番号・名称を抽出。条件リストで変換。"""
    if not opts:
        return ""

    raw_illust = ""

    # パターン1: "イラスト：【008.みつばち】" or "【イラスト】：007" or "【イラスト：014.いちご】"
    m = re.search(r"イラスト[】]?\s*[：:]\s*【?([^】\n]+?)】?\s*(?:名入れ|【|$|\n)", opts)
    if m:
        raw_illust = m.group(1).strip()
    # パターン1b: 【018.くま】【えぐちせしる】→ 先頭括弧がイラスト番号.名前
    if not raw_illust:
        m = re.match(r"\s*【(\d{2,3}[.\s][^】]+)】", opts)
        if m:
            raw_illust = m.group(1).strip()
    # パターン1c: カンマ区切りの先頭番号（016、漢字、ひらがな、ローマ字）
    if not raw_illust:
        m = re.match(r"\s*(\d{2,3})[、,]", opts)
        if m:
            raw_illust = m.group(1)
    # パターン2: 先頭の "028. やきゅう" (数字.イラスト名)
    if not raw_illust:
        m = re.match(r"\s*0*(\d+)\.\s*(\S+)", opts)
        if m:
            raw_illust = f"{m.group(1)}.{m.group(2)}"
    # パターン3: 改行後の「イラストスタンプ020」パターン
    if not raw_illust:
        m = re.search(r"イラスト(?:スタンプ)?[：:\s]*(\d+)", opts)
        if m:
            raw_illust = m.group(1)
    # パターン4: テキスト内「スタンプは004のひめ２で」
    if not raw_illust:
        m = re.search(r"スタンプ(?:は)?(\d+)の(\S+?)(?:で|$)", opts)
        if m:
            raw_illust = f"{m.group(1)}.{m.group(2)}"

    if not raw_illust:
        return ""

    # 条件リストで変換
    # "でんしゃ/025" → keyword="でんしゃ"
    m = re.match(r"(.+?)\s*/\s*(\d+)", raw_illust)
    if m:
        keyword = m.group(1).strip()
        result = _lookup_illust(keyword)
        if result:
            return result

    # "008. みつばち" → keyword="みつばち"
    m = re.match(r"0*(\d+)[.\s]+(.+)", raw_illust)
    if m:
        keyword = m.group(2).strip()
        result = _lookup_illust(keyword)
        if result:
            return result
        # 条件リストにない場合はそのまま2桁パディング
        return f"{int(m.group(1)):02d}.{keyword}"

    # 数字のみ → 条件リストの結果番号から逆引き
    m = re.match(r"0*(\d+)$", raw_illust)
    if m:
        num = int(m.group(1))
        # 全角→半角変換済みの数字
        target = f"{num:02d}."
        mapping = _get_illust_mapping()
        for mv in mapping.values():
            if mv.startswith(target):
                return mv
        return f"{num:02d}"

    # キーワードのみ
    result = _lookup_illust(raw_illust)
    return result if result else raw_illust


def _auto_column_width(ws, num_columns: int):
    """列幅を内容に合わせて自動調整"""
    for col_idx in range(1, num_columns + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value:
                    # 日本語文字は2倍幅で計算
                    val = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, char_len)
        # 最小8、最大40
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)


def _write_summary(wb: Workbook, summary_data: list):
    """Summaryシートを先頭に追加"""
    ws = wb.create_sheet(title="Summary", index=0)

    # ヘッダー
    headers = ["シート名", "件数", "数量合計"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # データ
    total_count = 0
    total_qty = 0
    for row_idx, item in enumerate(summary_data, 2):
        ws.cell(row=row_idx, column=1, value=item["シート名"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=item["件数"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=item["数量合計"]).border = THIN_BORDER
        total_count += item["件数"]
        total_qty += item["数量合計"]

    # 合計行
    total_row = len(summary_data) + 2
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total_count).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_qty).font = Font(bold=True)
    for col in range(1, 4):
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    # 列幅
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12


def generate_vendor_workbooks(df: pd.DataFrame) -> dict:
    """外注先別にExcelを生成して辞書で返す。

    Returns:
        dict[str, bytes]: {外注先名: Excelバイト列}
    """
    vendor_config = load_json("vendor_mapping.json")
    vendors = vendor_config.get("vendors", {})

    # 各行に出力シート名と製品カテゴリを付与
    df = df.copy()
    df["_出力シート"] = df.apply(_resolve_sheet_name, axis=1)

    results = {}

    for vendor_name, vendor_def in vendors.items():
        cats_rakuten = set(vendor_def.get("categories_rakuten", []))
        cats_yahoo = set(vendor_def.get("categories_yahoo", []))
        cats_amazon = set(vendor_def.get("categories_amazon", []))

        source_col = df.get("ソース", pd.Series(dtype=str))

        # 楽天: ソースが"rakuten"の行のみ
        mask_rakuten = (
            (source_col == "rakuten")
            & df["_出力シート"].isin(cats_rakuten)
        ) if cats_rakuten else pd.Series(False, index=df.index)

        # Yahoo!/Qoo10等: ソースが"non_rakuten"の行
        mask_yahoo = (
            (source_col == "non_rakuten")
            & df["_出力シート"].isin(cats_yahoo)
        ) if cats_yahoo else pd.Series(False, index=df.index)

        # 楽天+Yahoo!共通カテゴリ（categories_rakutenにある場合はnon_rakutenも含む）
        mask_rakuten_both = (
            (source_col.isin(["rakuten", "non_rakuten"]))
            & df["_出力シート"].isin(cats_rakuten)
        ) if cats_rakuten else pd.Series(False, index=df.index)

        # Amazon: 製品カテゴリでフィルタ
        mask_amazon = (
            (source_col == "amazon")
            & df["製品カテゴリ"].isin(cats_amazon)
        ) if cats_amazon else pd.Series(False, index=df.index)

        # 条件付きカテゴリ（単品/複数で外注先が変わる）
        mask_conditional = _build_conditional_mask(
            df, vendor_def.get("conditional_categories", {}), source_col
        )

        # メモキーワード（ひとことメモに特定キーワードを含む行）
        memo_keywords = vendor_def.get("memo_keywords", [])
        mask_memo = pd.Series(False, index=df.index)
        if memo_keywords and "ひとことメモ" in df.columns:
            memo_col = df["ひとことメモ"].fillna("")
            for kw in memo_keywords:
                mask_memo = mask_memo | memo_col.str.contains(kw, na=False)
            # 楽天/Yahoo!のみ
            mask_memo = mask_memo & source_col.isin(["rakuten", "non_rakuten"])

        vendor_df = df[mask_rakuten_both | mask_yahoo | mask_amazon | mask_conditional | mask_memo]

        if len(vendor_df) == 0:
            continue

        # 既存のgenerate_workbookを流用してExcelを生成
        wb_bytes = generate_workbook(vendor_df)
        results[vendor_name] = wb_bytes

    # 未割当データも出力（どの外注先にも属さない行）
    all_assigned = pd.Series(False, index=df.index)
    source_col = df.get("ソース", pd.Series(dtype=str))
    for vendor_name, vendor_def in vendors.items():
        cats_r = set(vendor_def.get("categories_rakuten", []))
        cats_y = set(vendor_def.get("categories_yahoo", []))
        cats_a = set(vendor_def.get("categories_amazon", []))
        mask_r = (
            source_col.isin(["rakuten", "non_rakuten"])
            & df["_出力シート"].isin(cats_r)
        ) if cats_r else pd.Series(False, index=df.index)
        mask_y = (
            (source_col == "non_rakuten")
            & df["_出力シート"].isin(cats_y)
        ) if cats_y else pd.Series(False, index=df.index)
        mask_a = (
            (source_col == "amazon")
            & df["製品カテゴリ"].isin(cats_a)
        ) if cats_a else pd.Series(False, index=df.index)
        mask_cond = _build_conditional_mask(
            df, vendor_def.get("conditional_categories", {}), source_col
        )
        memo_kws = vendor_def.get("memo_keywords", [])
        mask_memo = pd.Series(False, index=df.index)
        if memo_kws and "ひとことメモ" in df.columns:
            memo_c = df["ひとことメモ"].fillna("")
            for kw in memo_kws:
                mask_memo = mask_memo | memo_c.str.contains(kw, na=False)
            mask_memo = mask_memo & source_col.isin(["rakuten", "non_rakuten"])
        all_assigned = all_assigned | mask_r | mask_y | mask_a | mask_cond | mask_memo

    unassigned = df[~all_assigned]
    if len(unassigned) > 0:
        results["未割当"] = generate_workbook(unassigned)

    return results


def _build_conditional_mask(df, conditional_categories, source_col):
    """条件付きカテゴリのマスクを構築。

    conditional_categories例:
    {"おなまえスタンプ": {"単品複数": ["単品"]}}
    → _出力シートが「おなまえスタンプ」かつ単品複数が「単品」の行のみマッチ
    """
    mask = pd.Series(False, index=df.index)
    if not conditional_categories:
        return mask

    for sheet_name, conditions in conditional_categories.items():
        sheet_mask = df["_出力シート"] == sheet_name
        # 楽天/Yahoo!のみ（Amazonはcategories_amazonで別処理）
        sheet_mask = sheet_mask & source_col.isin(["rakuten", "non_rakuten"])
        for col_name, allowed_values in conditions.items():
            if col_name in df.columns:
                sheet_mask = sheet_mask & df[col_name].isin(allowed_values)
        mask = mask | sheet_mask

    return mask


def generate_vendor_zip(df: pd.DataFrame) -> bytes:
    """外注先別ExcelをZIPにまとめてバイト列で返す。"""
    import zipfile

    vendor_workbooks = generate_vendor_workbooks(df)
    today = datetime.now().strftime("%Y%m%d")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for vendor_name, wb_bytes in vendor_workbooks.items():
            filename = f"{vendor_name}_{today}.xlsx"
            zf.writestr(filename, wb_bytes)

    return output.getvalue()
