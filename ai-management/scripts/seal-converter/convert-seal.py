#!/usr/bin/env python3
"""
お名前シール変換スクリプト
3つのCSV（Amazon/Yahoo/楽天）→ お名前シールver5.3 Excel を自動生成する

使い方:
  python3 convert-seal.py <Amazon.csv> <Yahoo.csv> <楽天.csv> [--output 出力ファイル名.xlsx] [--date 0325]

引数:
  Amazon.csv  : Amazon受注CSV（例: 202603250814.csv）
  Yahoo.csv   : Yahoo受注CSV（例: 202603250829Y.csv）
  楽天.csv    : 楽天受注CSV（例: 202603250836R.csv）
  --output    : 出力ファイル名（デフォルト: お名前シール_MMDD.xlsx）
  --date      : 日付ラベル（デフォルト: 今日の日付）
  --template  : テンプレートファイルパス（お名前シールver5.3 .xlsm）
"""

import csv
import sys
import os
import argparse
import re
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("エラー: openpyxl が必要です。pip install openpyxl でインストールしてください。")
    sys.exit(1)


# === 定数 ===

# 出力Excelのヘッダー（23列）
HEADERS_FULL = [
    '注文日', '商品SKU', '商品コード', '受注番号', '商品名', '個数',
    '項目・選択肢', '備考', '注文者氏名', '受注ステータス', 'ひとことメモ',
    'GoQ管理番号(三桁ハイフン区切り)', 'バーコード', '複数', '単・複',
    'タイプ', 'タイプ2', 'タイプ3', '①印字', '②カット',
    '③内容', '④検品者', '⑤担当者'
]

# プチッと/ラミネート/Amazonは16列 or 13列
HEADERS_PUCHITO = [
    '注文日', '商品SKU', '商品SKU', '受注番号', '商品名', '個数',
    '項目・選択肢', '備考', '注文者氏名', '受注ステータス', 'ひとことメモ',
    'GoQ管理番号(三桁ハイフン区切り)', '複数', '', '', ''
]

HEADERS_AMAZON = [
    '注文日', '商品SKU', '商品コード', '受注番号', '商品名', '個数',
    '項目・選択肢', '備考', '注文者氏名', '受注ステータス', 'ひとことメモ',
    'GoQ管理番号(カスタム)', '複数'
]

# シート振り分けルール（項目・選択肢のパターンで判定）
# ####X または ##●●●##X のプレフィックスで判定する
SHEET_RULES = {
    'J': 'プチッと',      # ####J: くつにプチッと
    'K': 'ラミネート',     # ####K: 絵合わせシール
    'C': 'タグ',          # ####C: ノンアイロンタグ（イラスト付）
    'D': 'タグ',          # ####D: ノンアイロンタグ（シンプル）
    'N': 'タグ',          # ####N: ノンアイロンタグ（超特大等）
    'A': 'RYノーマル',     # ####A: ノーマルシール（イラスト付）
    'B': 'RYノーマル',     # ####B: シンプルシール
    'E': 'RYノーマル',     # ####E: 算数セット
    'F': 'RYノーマル',     # ####F: その他ノーマル
    'G': 'RYノーマル',     # ####G: その他ノーマル
    'H': 'RYノーマル',     # ####H: その他ノーマル
    'I': 'RYノーマル',     # ####I: その他ノーマル
}

# SKU→シートのフォールバックマッピング（####パターンがない場合に使用）
SKU_SHEET_MAP = {
    'NAMESEAL-NOR': 'RYノーマル',
    'NAMESEAL-NOR-S-SET': 'RYノーマル',
    'NAMESEAL-NOR-SA4-SET': 'RYノーマル',
    'NAMESEAL-NON': 'タグ',
    'NAMESEAL-NON-D': 'タグ',
    'NAMESEAL-NON-tagu': 'タグ',
    'NAMESEAL-LAMINATE': 'ラミネート',
    'onamae-seal': 'RYノーマル',
}


def detect_encoding(filepath):
    """CSVのエンコーディングを自動検出する"""
    for enc in ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(1000)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'cp932'


def read_csv(filepath, channel):
    """CSVを読み込み、正規化された辞書リストを返す"""
    enc = detect_encoding(filepath)
    rows = []

    with open(filepath, 'r', encoding=enc, errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    print(f"  {channel}: {len(rows)}行読み込み ({enc})")
    return rows


def is_seal_item(row, channel):
    """お名前シール関連の注文かどうか判定する（行単位で厳密に判定）"""
    memo = row.get('ひとことメモ', '')
    sku = row.get('商品SKU', '')
    product = row.get('商品名', '')
    options = row.get('項目・選択肢', '') or row.get('備考', '')

    # 商品名で非シール商品を除外（同GoQの別商品を弾く）
    exclude_keywords = [
        'スタンプ', 'おなまえポン', 'ゴム印', 'マルチインク', 'スタンプパッド',
        'ステイズオン', 'ソルベント', '補充インク', '印鑑', 'はんこ',
        '鉛筆', 'えんぴつ', 'ペンペン', 'ボールペン', 'ジェットストリーム',
    ]
    for kw in exclude_keywords:
        if kw in product:
            return False

    # 項目・選択肢にシール系のコードがあるか（最も信頼性が高い）
    if re.search(r'####[A-Z]', options) or re.search(r'##●●●##[A-Z]', options):
        # シール系コード（A,B,C,D,E,F,J,K,M,N）があるか
        code = extract_sheet_code(options)
        if code and code in SHEET_RULES:
            return True

    # 商品名でシール関連を判定
    seal_keywords = ['お名前シール', 'おなまえシール', 'プチッとネーム', '絵合わせシール',
                     'ノンアイロン', '名前シール']
    if any(kw in product for kw in seal_keywords):
        return True

    # ひとことメモの先頭で判定
    memo_first = memo.split('\n')[0].strip()
    if memo_first.startswith('おなまえシール'):
        # 同GoQ内の非シール商品でもメモが共有されるため、商品名も確認
        if 'シール' in product or 'seal' in sku.lower():
            return True

    # SKUで判定（シール系SKUのみ）
    seal_skus = {
        'onamae-seal', 'onamae-seal001', 'onamae-seal002', 'onamae-seal-sale',
        'NAMESEAL-NOR', 'NAMESEAL-NON', 'NAMESEAL-NON-D', 'NAMESEAL-NON-tagu',
        'NAMESEAL-LAMINATE', 'NAMESEAL-NOR-S-SET', 'NAMESEAL-NOR-SA4-SET',
    }
    if sku in seal_skus:
        return True

    return False


def extract_sheet_code(options):
    """項目・選択肢からシート判定用コード（A/B/C/D等）を抽出する
    パターン: ####X... または ##●●●##X... のプレフィックスから X を取得"""
    if not options:
        return None
    # ####X パターン
    m = re.search(r'####([A-Z])', options)
    if m:
        return m.group(1)
    # ##●●●##X パターン（2個目以降の商品でプレフィックスが変わる場合）
    m = re.search(r'##●●●##([A-Z])', options)
    if m:
        return m.group(1)
    return None


def determine_sheet(row, channel):
    """注文データをどのシートに振り分けるか判定する"""
    options = row.get('項目・選択肢', '')
    product = row.get('商品名', '')
    sku = row.get('商品SKU', '') or row.get('商品コード', '')

    # 1. 商品名でプチッと/ラミネートを最優先判定（Amazonでもこちらが優先）
    if 'プチッと' in product:
        return 'プチッと'
    if '絵合わせ' in product or 'ラミネート' in product:
        return 'ラミネート'

    # 2. 項目・選択肢のコードパターンで判定（最も信頼性が高い）
    code = extract_sheet_code(options)
    if code and code in SHEET_RULES:
        return SHEET_RULES[code]

    # 3. 商品名でバックアップ判定
    if 'ノンアイロン' in product:
        return 'タグ'

    # 4. SKUでフォールバック判定（####パターンがない注文）
    if sku in SKU_SHEET_MAP:
        return SKU_SHEET_MAP[sku]

    # 5. Amazon判定（プチッと/ラミネートでなければAmazonノーマル）
    if channel == 'Amazon':
        return 'Amazonノーマル'

    # デフォルト
    return 'RYノーマル'


def normalize_row(row, channel):
    """チャネル別のCSV行を統一フォーマットに正規化する"""
    if channel == 'Amazon':
        # Amazon CSVは列構成が異なる
        goq_key = 'GoQ管理番号(カスタム)'
        return {
            '注文日': row.get('注文日', ''),
            '商品SKU': 'onamae-seal',
            '商品コード': 'onamae-seal',
            '受注番号': row.get('受注番号', ''),
            '商品名': row.get('商品名', ''),
            '個数': row.get('個数', '1'),
            '項目・選択肢': row.get('備考', '').split('\n')[0] if '####' in row.get('備考', '') else '',
            '備考': 'Amazon',
            '注文者氏名': row.get('送付先氏名', ''),
            '受注ステータス': '',
            'ひとことメモ': row.get('ひとことメモ', ''),
            'GoQ': row.get(goq_key, ''),
        }
    elif channel == 'Yahoo':
        goq_key = 'GoQ管理番号(カスタム)'
        return {
            '注文日': row.get('注文日', ''),
            '商品SKU': 'onamae-seal',
            '商品コード': 'onamae-seal',
            '受注番号': row.get('受注番号', ''),
            '商品名': row.get('商品名', ''),
            '個数': row.get('個数', '1'),
            '項目・選択肢': row.get('項目・選択肢', ''),
            '備考': row.get('備考', ''),
            '注文者氏名': row.get('注文者氏名', ''),
            '受注ステータス': row.get('受注ステータス', ''),
            'ひとことメモ': row.get('ひとことメモ', ''),
            'GoQ': row.get(goq_key, ''),
        }
    else:  # 楽天
        goq_key = 'GoQ管理番号(三桁ハイフン区切り)'
        return {
            '注文日': row.get('注文日', ''),
            '商品SKU': 'onamae-seal',
            '商品コード': 'onamae-seal',
            '受注番号': row.get('受注番号', ''),
            '商品名': row.get('商品名', ''),
            '個数': row.get('個数', '1'),
            '項目・選択肢': row.get('項目・選択肢', ''),
            '備考': row.get('備考', ''),
            '注文者氏名': row.get('注文者氏名', ''),
            '受注ステータス': row.get('受注ステータス', ''),
            'ひとことメモ': row.get('ひとことメモ', ''),
            'GoQ': row.get(goq_key, ''),
        }


def expand_quantity(norm_row):
    """個数が2以上の場合、1行ずつに展開する（個数を1にして複製）
    完成品では個数2以上→1行ずつに展開されているため"""
    try:
        qty = int(float(norm_row.get('個数', '1')))
    except (ValueError, TypeError):
        qty = 1

    if qty <= 1:
        return [norm_row]

    expanded = []
    for _ in range(qty):
        row_copy = dict(norm_row)
        row_copy['個数'] = '1'
        expanded.append(row_copy)
    return expanded


def calc_barcode(goq):
    """バーコード列: GoQ番号を *xxxx-xxx* 形式に"""
    if not goq:
        return ''
    return f'*{goq}*'


def calc_fukusu(memo):
    """複数列: ひとことメモから単品/複数を判定"""
    if '単品' in memo:
        return '単品＋'
    if '複数' in memo:
        return '複数'
    return '単品'


def calc_tan_fuku(memo):
    """単・複列: 単品フラグ（複数でない=1）"""
    if '単品' in memo:
        return 0
    return 1


def calc_type(product):
    """タイプ列: ノンアイロンでない=1"""
    return 0 if 'ノンアイロン' in product else 1


def calc_type2(product):
    """タイプ2列: 透明でない=1"""
    return 0 if '透明' in product else 1


def calc_type3(product):
    """タイプ3列: ラバータイプでない=1"""
    return 0 if 'ラバータイプ' in product else 1


def calc_insatsu(options):
    """①印字列: ピンセット付等を判定"""
    # 項目・選択肢にピンセットの記載があるか
    if 'ピンセット' in options:
        return 'ピンセット付'
    return ''


def build_full_row(norm_row):
    """正規化された行データから23列の完全な行を構築する"""
    goq = norm_row['GoQ']
    memo = norm_row['ひとことメモ']
    product = norm_row['商品名']
    options = norm_row['項目・選択肢']

    return [
        norm_row['注文日'],
        norm_row['商品SKU'],
        norm_row['商品コード'],
        norm_row['受注番号'],
        norm_row['商品名'],
        norm_row['個数'],
        norm_row['項目・選択肢'],
        norm_row['備考'],
        norm_row['注文者氏名'],
        norm_row['受注ステータス'],
        norm_row['ひとことメモ'],
        goq,
        calc_barcode(goq),
        calc_fukusu(memo),
        calc_tan_fuku(memo),
        calc_type(product),
        calc_type2(product),
        calc_type3(product),
        calc_insatsu(options),
        '',  # ②カット
        '',  # ③内容
        '',  # ④検品者
        '',  # ⑤担当者
    ]


def build_puchito_row(norm_row):
    """プチッと/ラミネート用の16列行を構築する"""
    goq = norm_row['GoQ']
    memo = norm_row['ひとことメモ']

    return [
        norm_row['注文日'],
        norm_row['商品SKU'],
        norm_row['商品SKU'],  # 商品SKU重複
        norm_row['受注番号'],
        norm_row['商品名'],
        norm_row['個数'],
        norm_row['項目・選択肢'],
        norm_row['備考'],
        norm_row['注文者氏名'],
        norm_row['受注ステータス'],
        norm_row['ひとことメモ'],
        goq,
        calc_fukusu(memo),
        '', '', '',
    ]


def build_amazon_row(norm_row):
    """Amazonノーマル用の13列行を構築する"""
    goq = norm_row['GoQ']
    memo = norm_row['ひとことメモ']

    return [
        norm_row['注文日'],
        norm_row['商品SKU'],
        norm_row['商品コード'],
        norm_row['受注番号'],
        norm_row['商品名'],
        norm_row['個数'],
        norm_row['項目・選択肢'],
        norm_row['備考'],
        norm_row['注文者氏名'],
        norm_row['受注ステータス'],
        norm_row['ひとことメモ'],
        goq,
        calc_fukusu(memo),
    ]


def write_sheet(ws, headers, data_rows):
    """シートにヘッダー＋データを書き込む"""
    # ヘッダー
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    # データ
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(r, c, val)


def add_code_sheet(wb):
    """「コード」シートを追加（数式テンプレート）"""
    ws = wb.create_sheet('コード', 0)

    # ヘッダー行（列M～W）
    code_headers = {
        13: 'バーコード', 14: '複数', 15: '単・複',
        16: 'タイプ', 17: 'タイプ2', 18: 'タイプ3',
        19: '①印字', 20: '②カット', 21: '③内容',
        22: '④検品者', 23: '⑤担当者'
    }
    for c, h in code_headers.items():
        ws.cell(1, c, h)

    # 数式行（参考用）
    ws.cell(2, 13, '=IF(A2="","","*"&L2&"*")')
    ws.cell(2, 14, '=IF(COUNTIF(K2,"*単品*")>0,"単品＋",IF(COUNTIF(K2,"*複数*")>0,"複数","単品"))')
    ws.cell(2, 15, '=IF(A2="","",IF(COUNTIF(K2,"*単品*")=0,1,0))')
    ws.cell(2, 16, '=IF(A2="","",IF(COUNTIF(E2,"*ノンアイロン*")=0,1,0))')
    ws.cell(2, 17, '=IF(A2="","",IF(COUNTIF(E2,"*透明*")=0,1,0))')
    ws.cell(2, 18, '=IF(A2="","",IF(COUNTIF(E2,"*ラバータイプ*")=0,1,0))')


def main():
    parser = argparse.ArgumentParser(description='お名前シール変換スクリプト')
    parser.add_argument('amazon_csv', help='Amazon受注CSV')
    parser.add_argument('yahoo_csv', help='Yahoo受注CSV')
    parser.add_argument('rakuten_csv', help='楽天受注CSV')
    parser.add_argument('--output', '-o', help='出力ファイル名', default=None)
    parser.add_argument('--date', '-d', help='日付ラベル（例: 0325）', default=None)
    args = parser.parse_args()

    date_label = args.date or datetime.now().strftime('%m%d')
    output_path = args.output or f'お名前シール_{date_label}.xlsx'

    print(f"\n=== お名前シール変換 ({date_label}) ===\n")

    # 1. CSV読み込み
    print("【1】CSV読み込み")
    amazon_rows = read_csv(args.amazon_csv, 'Amazon')
    yahoo_rows = read_csv(args.yahoo_csv, 'Yahoo')
    rakuten_rows = read_csv(args.rakuten_csv, '楽天')

    # 2. お名前シールをフィルタ＆正規化
    print("\n【2】お名前シールフィルタリング")
    sheet_data = defaultdict(list)  # シート名 -> [行データ]
    all_data = []  # シート「1」用（全データ）

    expanded_count = 0
    for channel, rows in [('Amazon', amazon_rows), ('Yahoo', yahoo_rows), ('楽天', rakuten_rows)]:
        seal_count = 0
        for row in rows:
            if not is_seal_item(row, channel):
                continue
            seal_count += 1

            norm = normalize_row(row, channel)
            sheet_name = determine_sheet(row, channel)

            # 個数展開（個数2以上→1行ずつに展開）
            expanded_rows = expand_quantity(norm)
            if len(expanded_rows) > 1:
                expanded_count += len(expanded_rows) - 1

            for exp_norm in expanded_rows:
                # 全データ用
                full_row = build_full_row(exp_norm)
                all_data.append(full_row)

                # シート別データ
                if sheet_name in ('プチッと', 'ラミネート'):
                    sheet_data[sheet_name].append(build_puchito_row(exp_norm))
                elif sheet_name == 'Amazonノーマル':
                    sheet_data[sheet_name].append(build_amazon_row(exp_norm))
                else:
                    sheet_data[sheet_name].append(full_row)

        print(f"  {channel}: {seal_count}件抽出")

    if expanded_count:
        print(f"  ※個数展開: +{expanded_count}行（個数2以上を1行ずつに展開）")

    # 3. 出力
    print(f"\n【3】Excel出力")
    wb = openpyxl.Workbook()

    # コードシート
    add_code_sheet(wb)

    # シート「1」（全データ）
    ws1 = wb.create_sheet('1')
    write_sheet(ws1, HEADERS_FULL, all_data)
    print(f"  シート「1」: {len(all_data)}件")

    # RYノーマル
    ws_ry = wb.create_sheet('RYノーマル')
    ry_data = sheet_data.get('RYノーマル', [])
    write_sheet(ws_ry, HEADERS_FULL, ry_data)
    print(f"  シート「RYノーマル」: {len(ry_data)}件")

    # タグ
    ws_tag = wb.create_sheet('タグ')
    tag_data = sheet_data.get('タグ', [])
    write_sheet(ws_tag, HEADERS_FULL, tag_data)
    print(f"  シート「タグ」: {len(tag_data)}件")

    # プチッと
    ws_puchi = wb.create_sheet('プチッと')
    puchi_data = sheet_data.get('プチッと', [])
    write_sheet(ws_puchi, HEADERS_PUCHITO, puchi_data)
    print(f"  シート「プチッと」: {len(puchi_data)}件")

    # ラミネート
    ws_lami = wb.create_sheet('ラミネート')
    lami_data = sheet_data.get('ラミネート', [])
    write_sheet(ws_lami, HEADERS_PUCHITO, lami_data)
    print(f"  シート「ラミネート」: {len(lami_data)}件")

    # Amazonノーマル
    ws_amz = wb.create_sheet('Amazonノーマル')
    amz_data = sheet_data.get('Amazonノーマル', [])
    write_sheet(ws_amz, HEADERS_AMAZON, amz_data)
    print(f"  シート「Amazonノーマル」: {len(amz_data)}件")

    # デフォルトの空シートを削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(output_path)
    print(f"\n✓ 保存完了: {output_path}")

    # 4. サマリー
    print(f"\n{'='*50}")
    print(f"【サマリー】")
    print(f"  全データ: {len(all_data)}件")
    print(f"  RYノーマル: {len(ry_data)}件")
    print(f"  タグ: {len(tag_data)}件")
    print(f"  プチッと: {len(puchi_data)}件")
    print(f"  ラミネート: {len(lami_data)}件")
    print(f"  Amazonノーマル: {len(amz_data)}件")
    total_sheets = len(ry_data) + len(tag_data) + len(puchi_data) + len(lami_data) + len(amz_data)
    print(f"  シート合計: {total_sheets}件")
    if total_sheets != len(all_data):
        print(f"  ⚠ シート合計({total_sheets}) ≠ 全データ({len(all_data)}) — 差分確認してください")
    print(f"{'='*50}")


if __name__ == '__main__':
    main()
