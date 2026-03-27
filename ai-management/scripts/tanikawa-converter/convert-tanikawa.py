#!/usr/bin/env python3
"""
タニエバー割付表 自動変換スクリプト
外注マクロテンプレートの「タニエバー」シートから割付表を自動生成する

使い方:
  python3 convert-tanikawa.py <外注マクロテンプレート.xlsm> [--template 割付表テンプレート.xls] [--output 出力ファイル名.xlsx]

引数:
  外注マクロテンプレート.xlsm : 外注マクロテンプレートNEW のファイル
  --template : 割付表最新ver6.xls テンプレート（省略時はデフォルト構造で生成）
  --output   : 出力ファイル名（デフォルト: 割付表_タニエバー_自動生成.xlsx）
"""

import sys
import os
import re
import argparse
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("エラー: openpyxl が必要です。pip install openpyxl でインストールしてください。")
    sys.exit(1)


# === SKU→商品名マッピング ===
SKU_TO_PRODUCT = {
    'DOORNAME': '玄関ドア',
    'TSK-21688': 'ワンピース スタンペンG',
    'TSK-22616': 'ツインＧＴ',
    'TSK-22753': 'ツインＧＴ',
    'TSK-48876': 'ちいかわ スタンペンG',
    'TSK-48883': 'ちいかわ スタンペンG',
    'TSK-48890': 'ちいかわ スタンペンG',
    'TSK-50800': 'ちいかわ ツインGT キャップレス',
    'TSK-50817': 'ちいかわ ツインGT キャップレス',
    'TSK-54228': 'スタンペン4FCL',
    'TSK-56314': 'ツインＧＴロング',
    'TSK-65910': 'スタンペンG',
    'TSK-65927': 'スタンペンG',
    'TSK-68539': 'ネームＧキャップレス',
    'TSK-69116': 'ツインＧＴ キャップレス',
    'TSK-69123': 'ツインＧＴ キャップレス',
    'TSK-69130': 'ツインＧＴ キャップレス',
    'TSK-69147': 'ツインＧＴ キャップレス',
    'TSK-691HK': 'ツインＧＴ キャップレス',
    'rk-name': 'リラックマ ネームＧキャップレス',
    'st-g': 'スタンペンG',
    'st4fcl': 'スタンペン4FCL',
    'tani-gtk': 'GTK',
    'tani-n': 'ツインＧＴ キャップレス',
    'tani-p': 'スヌーピー スタンペンG',
    'tsk-48876': 'ちいかわ スタンペンG',
    'tuingt-01': 'ツインＧＴ キャップレス',
    'tuingt-02': 'ツインＧＴ',
    'tuingtgk': 'ツインＧＴ',
    'TSK-48753': 'スタンペン4FCL',
    'TSK-65903': 'スタンペンG',
    'TSK-67204': 'スタンペンG',
}

# === レイアウト変換テーブル ===
LAYOUT_MAP = {
    '姓のみ': '9mm',
    '姓+1文字': '9mm2文字+添え字',
    '姓+1文字(小さいもの)': '9mm2文字+添え字',
    '姓+1文字（小さいもの）': '9mm2文字+添え字',
    'フルネーム': '9mmフルネーム',
    '名前のみ': '9mm',
    '1文字': '9mm',
}


def resolve_product_name(sku):
    """SKUから商品名を解決する"""
    if not sku:
        return ''
    # 完全一致を先にチェック
    for prefix, name in sorted(SKU_TO_PRODUCT.items(), key=lambda x: -len(x[0])):
        if sku.startswith(prefix) or sku.lower().startswith(prefix.lower()):
            return name
    # ハイフン区切りの先頭2パートで再チェック
    parts = sku.split('-')
    for n in range(min(3, len(parts)), 0, -1):
        key = '-'.join(parts[:n])
        if key in SKU_TO_PRODUCT:
            return SKU_TO_PRODUCT[key]
        if key.lower() in {k.lower(): k for k in SKU_TO_PRODUCT}:
            for k, v in SKU_TO_PRODUCT.items():
                if k.lower() == key.lower():
                    return v
    return sku[:20]


def parse_options(options_text, memo_text=''):
    """項目・選択肢テキストをパースして各フィールドに分解する"""
    result = {
        'color': '',
        'font': '',
        'ink': '',
        'layout': '',
        'layout_9mm': '',
        'size': '',
        'name_9mm': '',
        'name_5mm': '',
        'name_10mm': '',
        'layout_5mm': '',
    }

    if not options_text:
        return result

    lines = [l.strip() for l in options_text.split('\n') if l.strip()]

    # 区切り文字を統一するヘルパー（":"と"="の両方に対応）
    def split_value(line):
        """キー:値 または キー=値 から値を抽出"""
        if ':' in line:
            return line.split(':', 1)[-1].strip()
        if '=' in line:
            return line.split('=', 1)[-1].strip()
        return ''

    for line in lines:
        # 不要行をスキップ
        if '営業日' in line or '(tani)' in line or 'tani)' in line:
            continue
        if '要確認' in line or 'キャンセル' in line:
            continue

        # カラー + サイズ（ツインGT系）"カラー + サイズ:..." or "カラー + サイズ=..."
        if line.startswith('カラー') and 'サイズ' in line:
            val = split_value(line)
            # "ブラック GK(9ミリ+10ミリ黒檀)" or "シルバー(9ミリ+5ミリ+10ミリ黒檀)"
            m = re.match(r'^(\S+?)[\s\(（]', val)
            if m:
                result['color'] = m.group(1)
            else:
                result['color'] = val.split()[0] if val else ''
            continue

        # インク色: "インク色:朱色", "インク色=朱色", "インク朱"
        if 'インク' in line and 'レイアウト' not in line and '作成名' not in line:
            val = split_value(line)
            # "色:朱色" → "朱色", "色=朱色" → "朱色"
            if val.startswith('色'):
                val = val[1:].lstrip('=: ').strip()
            if not val:
                # "インク朱" のようなパターン
                val = line.replace('インク', '').strip()
            # "朱" → "朱色" に正規化
            if val == '朱':
                val = '朱色'
            result['ink'] = val
            continue

        # 5mm/10mm レイアウト（9mmより先にチェック）
        if ('5mm' in line.lower() or '5ｍｍ' in line) and 'レイアウト' in line and '作成名' not in line:
            result['layout_5mm'] = split_value(line)
            continue

        # 10mm レイアウト
        if ('10mm' in line.lower() or '10ｍｍ' in line) and 'レイアウト' in line and '作成名' not in line:
            continue

        # 9mmレイアウト（5mm/10mmを除外済み）
        if 'レイアウト' in line and '作成名' not in line:
            layout_val = split_value(line)
            result['layout'] = layout_val
            for key, mapped in LAYOUT_MAP.items():
                if key in layout_val:
                    result['layout_9mm'] = mapped
                    break
            if not result['layout_9mm']:
                result['layout_9mm'] = '9mm'
            continue

        # 5mm/10mm 作成名（9mmより先にチェック）
        if ('5mm' in line.lower() or '5ｍｍ' in line) and ('作成名' in line or '丸作成名' in line):
            result['name_5mm'] = split_value(line)
            continue

        if ('10mm' in line.lower() or '10ｍｍ' in line) and ('作成名' in line or '丸作成名' in line):
            result['name_10mm'] = split_value(line)
            continue

        # 9mm作成名: "作成名(アルファベット不可):佐藤", "①作成名（...）:矢部", "作成名...=菅原"
        if '作成名' in line:
            val = split_value(line)
            # "①作成名..." のような番号プレフィックスを除去
            cleaned = re.sub(r'^[①②③]\s*', '', line)
            if not val:
                val = split_value(cleaned)
            result['name_9mm'] = val
            continue

        # 書体: "楷書体", "楷書体のみです", "5mm丸訂正印楷書体のみです"
        # 5mm丸訂正印の書体行はスキップ（別用途）
        if '書体' in line:
            if '5mm丸訂正印' in line or '5ｍｍ丸訂正印' in line:
                continue
            val = line.replace('のみです', '').strip()
            val = re.sub(r'^\d+mm浸透\s*', '', val)
            result['font'] = val
            continue

        # サイズ
        if line.startswith('サイズ') and ('=' in line or ':' in line):
            result['size'] = split_value(line)
            continue

        # 残り: カラーの可能性（GT(9ミリ+5ミリ)等のサイズ情報を除去）
        if not result['color']:
            colors = ['ブラック', 'シルバー', 'ブルー', 'ブラウン', 'ピンク', 'レッド',
                      'ホワイト', 'ネイビー', 'グリーン', 'パープル', 'ゴールド',
                      'アイボリー', 'スケルトン', 'ハンコヤ', 'イエロー', 'ゾロ']
            if any(c in line for c in colors):
                # カラー名だけ抽出（"ブラック　GT(9ミリ+5ミリ)" → "ブラック"）
                m = re.match(r'^(\S+?)[\s　]+(GT|GK)', line)
                if m:
                    result['color'] = m.group(1)
                else:
                    # "スケルトンピンク" → "ピンク" 等の短縮（先頭の修飾語を除去）
                    result['color'] = line
                continue

        # 書体が単独行で来るパターン（"丸ゴシック体", "ポップ体", "古印体"等）
        known_fonts = ['楷書体', '行書体', '古印体', '明朝体', '丸ゴシック体', 'ポップ体',
                       'ゴシック体', '隷書体', '印相体', '篆書体']
        if any(f in line for f in known_fonts) and not result['font']:
            result['font'] = line.replace('のみです', '').strip()
            continue

        # 書体なし・作成名なしで残っている行→作成名の可能性（ワンピース等の特殊形式）
        if not result['name_9mm'] and not any(kw in line for kw in ['レイアウト', 'インク', '書体', 'サイズ', '営業', 'カラー', '丸訂正印']):
            result['name_9mm'] = line
            continue

    # デフォルト値
    if not result['ink']:
        result['ink'] = '朱色'
    if not result['layout_9mm']:
        result['layout_9mm'] = '9mm'

    return result


def detect_memo_extras(sku, options_text, memo_text=''):
    """SKUや項目から追加の備考情報を抽出する"""
    extras = []
    if '黒檀' in (options_text or '') or '黒壇' in (options_text or ''):
        extras.append('黒檀')
    elif '黒檀' in sku:
        extras.append('黒檀')
    return '\n'.join(extras) if extras else ''


def fix_gk_names(item, options_text):
    """GK(9mm+10mm)タイプの場合、「5mmまたは10mm丸作成名」は10mmに振り分ける"""
    if not options_text:
        return
    # GKタイプ判定: "GK(9ミリ+10ミリ" が含まれる場合
    is_gk = 'GK' in options_text and '10ミリ' in options_text
    if is_gk:
        # 「5mmまたは10mm丸作成名」→ 10mm作成名として扱う
        if item['name_5mm'] and not item['name_10mm']:
            item['name_10mm'] = item['name_5mm']
            item['name_5mm'] = ''
    # GTKタイプ: 9mm+5mm+10mm 全てに同じ名前
    if 'GTK' in options_text or ('9ミリ' in options_text and '5ミリ' in options_text and '10ミリ' in options_text):
        # 名前が1つしかない場合、全サイズに展開
        name = item['name_9mm']
        if name and not item['name_5mm']:
            item['name_5mm'] = name
        if name and not item['name_10mm']:
            item['name_10mm'] = name


def convert_tanikawa(macro_path, output_path, template_path=None):
    """外注マクロテンプレートからタニエバー割付表を生成する"""
    print(f"入力: {macro_path}")

    # 外注マクロテンプレート読み込み
    wb_in = openpyxl.load_workbook(macro_path, data_only=True)
    if 'タニエバー' not in wb_in.sheetnames:
        print("エラー: 「タニエバー」シートが見つかりません")
        sys.exit(1)

    ws_in = wb_in['タニエバー']
    rows_in = list(ws_in.iter_rows(values_only=True))
    hdr_in = [str(c) if c else '' for c in rows_in[0]]

    # 列インデックス
    idx = {h: i for i, h in enumerate(hdr_in)}

    # データ読み込み
    items = []
    for row in rows_in[1:]:
        vals = {h: (str(row[i]).strip() if i < len(row) and row[i] else '') for i, h in enumerate(hdr_in)}
        goq = vals.get('GoQ管理番号(三桁ハイフン区切り)', '')
        if not goq:
            continue

        sku = vals.get('商品SKU', '')
        options = vals.get('項目・選択肢', '')
        memo = vals.get('備考', '')
        buyer = vals.get('注文者氏名', '')
        qty = vals.get('個数', '1')
        fukusu = vals.get('複数', '')
        order_no = vals.get('受注番号', '')
        sku_code = vals.get('商品コード', '') or vals.get('商品コード2', '')
        hitokoto = vals.get('ひとことメモ', '')

        # 商品名解決
        product_name = resolve_product_name(sku)

        # 項目・選択肢パース
        parsed = parse_options(options, memo)

        # 備考の追加情報
        extra_memo = detect_memo_extras(sku, options, memo)
        final_memo = '\n'.join(filter(None, [extra_memo, memo])) if extra_memo or memo else ''

        # GK/GTKタイプの5mm/10mm作成名修正
        fix_gk_names(parsed, options)

        # GTK特殊処理: "10mm丸黒檀レイアウト=姓のみ(古印体)" から書体・作成名を抽出
        for line in options.split('\n'):
            line = line.strip()
            if '10mm丸黒檀' in line and 'レイアウト' in line:
                m = re.search(r'[\(（](.+?)[\)）]', line)
                if m and not parsed['font']:
                    parsed['font'] = m.group(1)

        # 作成名が空の場合、注文者氏名の姓をフォールバック
        if not parsed['name_9mm'] and buyer:
            parsed['name_9mm'] = buyer.split()[0] if ' ' in buyer else buyer.split('　')[0] if '　' in buyer else buyer

        # GTK/GKで作成名があるが5mm/10mmが空の場合、注文者氏名の姓でフォールバック
        if parsed['name_9mm'] and not parsed['name_5mm'] and not parsed['name_10mm']:
            # GTKの場合は5mmと10mmにも展開
            if 'GTK' in options or ('9ミリ' in options and '5ミリ' in options and '10ミリ' in options):
                parsed['name_5mm'] = parsed['name_9mm']
                parsed['name_10mm'] = parsed['name_9mm']

        # カラー名の正規化（修飾語を短縮）
        color_map = {
            'スケルトンピンク': 'ピンク',
            'スケルトンブルー': 'ブルー',
            'スケルトングリーン': 'グリーン',
        }
        if parsed['color'] in color_map:
            parsed['color'] = color_map[parsed['color']]

        items.append({
            'goq': goq,
            'sku': sku,
            'sku_code': sku_code,
            'order_no': order_no,
            'product_name': product_name,
            'color': parsed['color'],
            'font': parsed['font'],
            'ink': parsed['ink'],
            'layout_9mm': parsed['layout_9mm'],
            'size': parsed['size'],
            'name_9mm': parsed['name_9mm'],
            'name_5mm': parsed['name_5mm'],
            'name_10mm': parsed['name_10mm'],
            'qty': qty,
            'memo': final_memo,
            'buyer': buyer,
            'fukusu': fukusu,
            'options_raw': options,
            'hitokoto': hitokoto,
        })

    print(f"  タニエバー: {len(items)}件読み込み")
    wb_in.close()

    # 個数2以上の展開（1行ずつに分割、個数は1に）
    # ただし同GoQで既に複数行ある場合は展開しない（既に分割済み）
    goq_counts = defaultdict(int)
    for item in items:
        goq_counts[item['goq']] += 1

    expanded = []
    for item in items:
        qty = int(float(item['qty'])) if item['qty'] else 1
        # 同GoQに1行しかなく、個数>1の場合のみ展開
        if qty > 1 and goq_counts[item['goq']] == 1:
            for _ in range(qty):
                new_item = dict(item)
                new_item['qty'] = '1'
                expanded.append(new_item)
        else:
            item['qty'] = '1'
            expanded.append(item)
    if len(expanded) != len(items):
        print(f"  個数展開: {len(items)}→{len(expanded)}行")
    items = expanded

    # === 出力Excel生成 ===
    wb_out = openpyxl.Workbook()

    # --- シート「1」（メイン割付表）---
    ws1 = wb_out.active
    ws1.title = '1'
    headers_1 = ['番号', '商品名', 'カラー', '書体', 'インク', '9㎜丸レイアウト',
                 'サイズ', '9㎜', '5㎜', '10㎜', '個数', '備考', 'バーコード', '注文者氏名', '複数']

    # ヘッダー書き込み
    header_font = Font(bold=True)
    for c, h in enumerate(headers_1, 1):
        cell = ws1.cell(1, c, h)
        cell.font = header_font

    for i, item in enumerate(items, 1):
        ws1.cell(i + 1, 1, i)  # 番号
        ws1.cell(i + 1, 2, item['product_name'])
        ws1.cell(i + 1, 3, item['color'])
        ws1.cell(i + 1, 4, item['font'])
        ws1.cell(i + 1, 5, item['ink'])
        ws1.cell(i + 1, 6, item['layout_9mm'])
        ws1.cell(i + 1, 7, item['size'])
        ws1.cell(i + 1, 8, item['name_9mm'])
        ws1.cell(i + 1, 9, item['name_5mm'])
        ws1.cell(i + 1, 10, item['name_10mm'])
        ws1.cell(i + 1, 11, int(float(item['qty'])) if item['qty'] else 1)
        ws1.cell(i + 1, 12, item['memo'])
        ws1.cell(i + 1, 13, item['goq'])
        ws1.cell(i + 1, 14, item['buyer'])
        ws1.cell(i + 1, 15, item['fukusu'])

    # --- シート「9mm」---
    ws_9mm = wb_out.create_sheet('9ｍｍ')
    headers_9mm = ['商品名', '作成名', '書体', '数量', '備考', '確認']
    for c, h in enumerate(headers_9mm, 1):
        ws_9mm.cell(1, c, h).font = header_font

    row_9mm = 2
    for item in items:
        if item['name_9mm']:
            ws_9mm.cell(row_9mm, 1, item['layout_9mm'] or '9mm')
            ws_9mm.cell(row_9mm, 2, item['name_9mm'])
            ws_9mm.cell(row_9mm, 3, item['font'])
            ws_9mm.cell(row_9mm, 4, int(float(item['qty'])) if item['qty'] else 1)
            ws_9mm.cell(row_9mm, 5, item['memo'])
            row_9mm += 1

    # --- シート「5mm」---
    ws_5mm = wb_out.create_sheet('5ｍｍ')
    for c, h in enumerate(headers_9mm, 1):
        ws_5mm.cell(1, c, h).font = header_font

    row_5mm = 2
    for item in items:
        if item['name_5mm']:
            ws_5mm.cell(row_5mm, 1, '5mm')
            ws_5mm.cell(row_5mm, 2, item['name_5mm'])
            ws_5mm.cell(row_5mm, 3, item['font'])
            ws_5mm.cell(row_5mm, 4, int(float(item['qty'])) if item['qty'] else 1)
            ws_5mm.cell(row_5mm, 5, item['memo'])
            row_5mm += 1

    # --- シート「印刷用」（フル情報）---
    ws_print = wb_out.create_sheet('印刷用')
    headers_print = ['番号', '商品SKU', '商品コード2', '商品コード', '受注番号',
                     '商品名', '個数', 'カラー', '書体', 'インク',
                     '項目・選択肢', '項目・選択肢', '項目・選択肢', '項目・選択肢',
                     '項目・選択肢', '項目・選択肢',
                     '9㎜丸レイアウト', 'サイズ', '9㎜', '5㎜',
                     '10㎜', '個数', '備考', 'バーコード', '注文者氏名',
                     '受注ステータス', 'ひとことメモ', '複数']
    for c, h in enumerate(headers_print, 1):
        ws_print.cell(1, c, h).font = header_font

    for i, item in enumerate(items, 1):
        ws_print.cell(i + 1, 1, i)
        ws_print.cell(i + 1, 2, item['sku'])
        ws_print.cell(i + 1, 3, item['sku_code'])
        ws_print.cell(i + 1, 4, item['sku_code'])
        ws_print.cell(i + 1, 5, item['order_no'])
        ws_print.cell(i + 1, 6, item['product_name'])
        ws_print.cell(i + 1, 7, int(float(item['qty'])) if item['qty'] else 1)
        ws_print.cell(i + 1, 8, item['color'])
        ws_print.cell(i + 1, 9, item['font'])
        ws_print.cell(i + 1, 10, item['ink'])
        # 項目・選択肢は元データの各行を展開
        opt_lines = item['options_raw'].split('\n') if item['options_raw'] else []
        for j in range(6):
            ws_print.cell(i + 1, 11 + j, opt_lines[j].strip() if j < len(opt_lines) else '')
        ws_print.cell(i + 1, 17, item['layout_9mm'])
        ws_print.cell(i + 1, 18, item['size'])
        ws_print.cell(i + 1, 19, item['name_9mm'])
        ws_print.cell(i + 1, 20, item['name_5mm'])
        ws_print.cell(i + 1, 21, item['name_10mm'])
        ws_print.cell(i + 1, 22, int(float(item['qty'])) if item['qty'] else 1)
        ws_print.cell(i + 1, 23, item['memo'])
        ws_print.cell(i + 1, 24, item['goq'])
        ws_print.cell(i + 1, 25, item['buyer'])
        ws_print.cell(i + 1, 26, '')  # 受注ステータス
        ws_print.cell(i + 1, 27, item['hitokoto'])
        ws_print.cell(i + 1, 28, item['fukusu'])

    # --- Sheet2（レイアウト変換テーブル）---
    ws_s2 = wb_out.create_sheet('Sheet2')
    layout_table = [
        ['レイアウト=姓のみ', '9mm', '姓のみ', '5mm'],
        ['レイアウト=姓+1文字', '9mm2文字+添え字', '', ''],
        ['レイアウト=フルネーム', '9mmフルネーム', '', ''],
        ['レイアウト:フルネーム', '9mmフルネーム', '', ''],
        ['レイアウト:姓+1文字', '9mm2文字+添え字', '', ''],
        ['レイアウト:姓のみ', '9mm', '', ''],
        ['レイアウト:名前のみ', '9mm', '', ''],
        ['レイアウト=姓+1文字(小さいもの)', '9mm2文字+添え字', '', ''],
        ['レイアウト:姓+1文字（小さいもの）', '9mm2文字+添え字', '', ''],
        ['レイアウト=名前のみ', '9mm', '', ''],
    ]
    for r, row in enumerate(layout_table, 1):
        for c, val in enumerate(row, 1):
            ws_s2.cell(r, c, val)

    # 保存
    wb_out.save(output_path)
    print(f"\n出力: {output_path}")
    print(f"  シート「1」: {len(items)}行")
    print(f"  シート「9mm」: {row_9mm - 2}行")
    print(f"  シート「5mm」: {row_5mm - 2}行")
    print(f"  シート「印刷用」: {len(items)}行")


def main():
    parser = argparse.ArgumentParser(description='タニエバー割付表自動変換')
    parser.add_argument('input', help='外注マクロテンプレートファイル (.xlsm)')
    parser.add_argument('--template', help='割付表テンプレートファイル (.xls)')
    parser.add_argument('--output', help='出力ファイル名', default=None)
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"エラー: ファイルが見つかりません: {args.input}")
        sys.exit(1)

    output = args.output or '割付表_タニエバー_自動生成.xlsx'
    convert_tanikawa(args.input, output, args.template)


if __name__ == '__main__':
    main()
