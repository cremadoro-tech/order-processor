#!/usr/bin/env python3
"""
筆記具割付表 自動変換スクリプト
3つのCSV（楽天/Yahoo/Amazon）→ 割付表Excel を自動生成する

使い方:
  python3 convert-pen.py <楽天.csv> <Yahoo.csv> <Amazon.csv> --master <③割付表ver5.xls> --db <①振り分け表.xls> [--output 出力ファイル名.xlsx]
"""

import csv
import sys
import os
import re
import json
import argparse
from collections import defaultdict
from unicodedata import normalize

try:
    import openpyxl
    import xlrd
    from openpyxl.styles import Font
except ImportError:
    print("エラー: openpyxl, xlrd が必要です。pip install openpyxl xlrd でインストールしてください。")
    sys.exit(1)


# === Sheet1 ヘッダー（22列）===
HEADERS = [
    '注文日', '商品SKU', '商品コード', '受注番号', '商品名', '個数',
    'ボディーカラー', '文字カラー', '型式', '書体', '作成名', '作成名変換',
    '文字数', 'イラスト', 'テンプレート名', '備考', '注文者氏名',
    '受注ステータス', 'ひとことメモ', 'GoQ管理番号(三桁ハイフン区切り)',
    'バーコード', '複数'
]


# === マスターデータ読み込み ===

def load_master_data(master_path):
    """③割付表ver5からマスターデータを読み込む"""
    wb = xlrd.open_workbook(master_path)
    master = {}

    # 文字カラーマスター: ボディーカラー→文字カラー
    sheet = wb.sheet_by_name('文字カラー')
    color_map = {}
    for r in range(sheet.nrows):
        body = str(sheet.cell_value(r, 0)).strip()
        text = str(sheet.cell_value(r, 1)).strip()
        if body and text:
            color_map[body] = text
    master['color_map'] = color_map

    # テンプレートマスター: 商品コード→テンプレート名
    sheet = wb.sheet_by_name('テンプレート')
    template_map = {}
    for r in range(sheet.nrows):
        code = str(sheet.cell_value(r, 0)).strip()
        tmpl = str(sheet.cell_value(r, 1)).strip()
        if code and tmpl:
            template_map[code] = tmpl
    master['template_map'] = template_map

    # 型式マスター: 商品コード→型式
    sheet = wb.sheet_by_name('型式')
    type_map = {}
    for r in range(sheet.nrows):
        code = str(sheet.cell_value(r, 0)).strip()
        ttype = str(sheet.cell_value(r, 1)).strip()
        if code:
            type_map[code] = ttype
    master['type_map'] = type_map

    # イラスト変換マスター: (イラスト, 文字カラー)→テンプレート名
    sheet = wb.sheet_by_name('イラスト変換')
    illust_map = {}
    for r in range(sheet.nrows):
        illust = str(sheet.cell_value(r, 0)).strip()
        # 列0=イラスト, 列1=文字カラー, 列2=テンプレート名（20文字以下版）
        # 列4=文字カラー結合キー, 列5=テンプレート名（短縮版）
        for col_color, col_tmpl in [(1, 5), (1, 2)]:
            color = str(sheet.cell_value(r, col_color)).strip() if sheet.ncols > col_color else ''
            tmpl = str(sheet.cell_value(r, col_tmpl)).strip() if sheet.ncols > col_tmpl else ''
            if illust and color and tmpl:
                illust_map[f"{illust}|{color}"] = tmpl
    master['illust_map'] = illust_map

    return master


def load_product_db(db_path):
    """①ジェットストリーム振り分け表からマッピングDBを読み込む"""
    wb = xlrd.open_workbook(db_path)
    sheet = wb.sheet_by_name('データベース')
    product_map = {}
    for r in range(sheet.nrows):
        code = str(sheet.cell_value(r, 0)).strip()
        code2 = str(sheet.cell_value(r, 1)).strip()
        category = str(sheet.cell_value(r, 2)).strip()
        if code:
            product_map[code] = {'code2': code2, 'category': category}
    return product_map


# === CSV読み込み ===

def detect_encoding(filepath):
    """CSVのエンコーディングを自動検出"""
    for enc in ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(2000)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'cp932'


def read_csv_rows(filepath, channel):
    """CSVを読み込み、チャネル情報を付加した辞書リストを返す"""
    enc = detect_encoding(filepath)
    rows = []
    with open(filepath, 'r', encoding=enc, errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            row['_channel'] = channel
            rows.append(row)
    print(f"  {channel}: {len(rows)}行読み込み")
    return rows


def normalize_columns(rows, channel):
    """列名を統一する"""
    normalized = []
    for row in rows:
        n = {}
        if channel == 'Amazon':
            n['注文日'] = row.get('注文日', '')
            n['商品SKU'] = row.get('商品SKU', '')
            n['商品コード'] = row.get('商品SKU', '')
            n['受注番号'] = row.get('受注番号', '')
            n['商品名'] = row.get('商品名', '')
            n['個数'] = row.get('個数', '1')
            n['項目・選択肢'] = ''
            n['備考'] = row.get('備考', '')
            n['注文者氏名'] = row.get('送付先氏名', '')
            n['受注ステータス'] = ''
            n['ひとことメモ'] = row.get('ひとことメモ', '')
            n['GoQ管理番号'] = row.get('GoQ管理番号(カスタム)', '')
        else:
            n['注文日'] = row.get('注文日', '')
            n['商品SKU'] = row.get('商品SKU', '')
            n['商品コード'] = row.get('商品コード', row.get('商品SKU', ''))
            n['受注番号'] = row.get('受注番号', '')
            n['商品名'] = row.get('商品名', '')
            n['個数'] = row.get('個数', '1')
            n['項目・選択肢'] = row.get('項目・選択肢', '')
            n['備考'] = row.get('備考', '')
            n['注文者氏名'] = row.get('注文者氏名', '')
            n['受注ステータス'] = row.get('受注ステータス', '')
            n['ひとことメモ'] = row.get('ひとことメモ', '')
            n['GoQ管理番号'] = row.get('GoQ管理番号(三桁ハイフン区切り)',
                                    row.get('GoQ管理番号(カスタム)', ''))
        n['_channel'] = channel
        normalized.append(n)
    return normalized


# === フィルタリング ===

def is_pen_item(row, product_db):
    """筆記具関連の注文かどうか判定"""
    memo = row.get('ひとことメモ', '')
    memo_first = memo.split('\n')[0].strip()

    # ひとことメモで判定
    pen_keywords = ['ジェットストリーム', 'クルトガ', 'ラミー']
    if any(kw in memo_first for kw in pen_keywords):
        return True

    # 商品コードでDB照合
    code = row.get('商品コード', '')
    if code in product_db:
        cat = product_db[code].get('category', '')
        if cat == 'ジェットストリーム':
            return True

    return False


# === 項目・選択肢パース ===

def parse_options(options_text):
    """項目・選択肢テキストをパースしてフィールドを抽出"""
    result = {
        'body_color': '',
        'font': '',
        'name': '',
        'illustration': '',
        'message_card': '',
    }

    if not options_text:
        return result

    lines = [l.strip() for l in options_text.split('\n') if l.strip()]

    for line in lines:
        # 不要行スキップ
        if '営業日' in line or '要確認' in line or '了承した' in line:
            continue
        if '転売防止' in line or '配送方法' in line:
            continue

        # ボディーカラー
        if '本体カラー' in line or 'カラー' in line:
            val = _extract_value(line)
            if val and not result['body_color']:
                result['body_color'] = val
            continue

        # 書体
        if '書体' in line:
            val = _extract_value(line)
            if val:
                result['font'] = val
            continue

        # 作成名（名入れ内容）
        if '名入れ内容' in line or '作成名' in line:
            val = _extract_value(line)
            if val:
                result['name'] = val
            continue

        # イラスト
        if 'イラスト' in line and '選択' not in line:
            val = _extract_value(line)
            if val:
                result['illustration'] = val
            continue

        # メッセージカード
        if 'メッセージカード' in line:
            val = _extract_value(line)
            if val:
                result['message_card'] = val
            continue

        # 名入れ文字（ひらがな/カタカナ/英数字等）→ スキップ
        if '名入れ文字' in line or '名入れ:' in line:
            continue

    return result


def _extract_value(line):
    """行から値部分を抽出（キー:値 or キー=値）"""
    if ':' in line:
        return line.split(':', 1)[-1].strip()
    if '=' in line:
        return line.split('=', 1)[-1].strip()
    return ''


# === 変換処理 ===

def zenkaku_to_hankaku(text):
    """全角英数→半角に変換"""
    if not text:
        return ''
    return normalize('NFKC', text).replace('\u3000', ' ').strip()


def determine_quantity_type(memo):
    """ひとことメモから単品/複数/単品+を判定"""
    if not memo:
        return '単品'
    lines = memo.split('\n')
    # 複数行にわたって判定
    has_fukusu = any('複数' in l for l in lines)
    has_tanpin = any('単品' in l for l in lines)

    if has_tanpin and has_fukusu:
        return '単品+'
    if has_fukusu:
        return '複数'
    return '単品'


def lookup_master(value, master_dict, fuzzy=False):
    """マスターデータから値を検索（完全一致→前方一致）"""
    if not value:
        return ''
    if value in master_dict:
        return master_dict[value]
    if fuzzy:
        # ハイフン区切りで末尾を削って検索
        parts = value.split('-')
        for i in range(len(parts) - 1, 0, -1):
            prefix = '-'.join(parts[:i])
            if prefix in master_dict:
                return master_dict[prefix]
    return ''


def convert_row(row, master, product_db):
    """1行のCSVデータを22列の割付表行に変換"""
    code = row.get('商品コード', '')
    options = row.get('項目・選択肢', '')
    memo = row.get('ひとことメモ', '')
    goq = row.get('GoQ管理番号', '')

    # 項目選択肢パース
    parsed = parse_options(options)

    # ボディーカラー
    body_color = parsed['body_color']
    # SKUにカラー情報が含まれる場合のフォールバック
    sku = row.get('商品SKU', '')

    # 文字カラー（マスター参照）
    text_color = lookup_master(body_color, master['color_map'])

    # 型式（マスター参照）
    katashiki = lookup_master(code, master['type_map'], fuzzy=True)

    # 書体
    font = parsed['font']

    # 作成名
    name_raw = parsed['name']
    name_converted = zenkaku_to_hankaku(name_raw)
    char_count = len(name_converted) if name_converted else ''

    # イラスト
    illustration = parsed['illustration']

    # テンプレート名
    template = ''
    if illustration and text_color:
        # イラスト変換マスター参照
        key = f"{illustration}|{text_color}"
        template = master['illust_map'].get(key, '')
        # 文字数制限付きバージョンも試す
        if not template:
            for suffix in ['20文字以下', '20文字', '6文字', '4文字', '1文字']:
                key_with_len = f"{illustration}|{text_color}"
                if key_with_len in master['illust_map']:
                    template = master['illust_map'][key_with_len]
                    break
    if not template:
        # テンプレートマスター参照
        template = lookup_master(code, master['template_map'], fuzzy=True)

    # 備考
    remark = parsed.get('message_card', '') or row.get('備考', '')

    # 複数判定
    qty_type = determine_quantity_type(memo)

    # バーコード
    barcode = f"*{goq}*" if goq else ''

    return {
        '注文日': row.get('注文日', ''),
        '商品SKU': sku,
        '商品コード': code,
        '受注番号': row.get('受注番号', ''),
        '商品名': row.get('商品名', ''),
        '個数': row.get('個数', '1'),
        'ボディーカラー': body_color,
        '文字カラー': text_color,
        '型式': katashiki,
        '書体': font,
        '作成名': name_raw,
        '作成名変換': name_converted,
        '文字数': char_count,
        'イラスト': illustration,
        'テンプレート名': template,
        '備考': remark,
        '注文者氏名': row.get('注文者氏名', ''),
        '受注ステータス': row.get('受注ステータス', ''),
        'ひとことメモ': memo,
        'GoQ管理番号(三桁ハイフン区切り)': goq,
        'バーコード': barcode,
        '複数': qty_type,
    }


# === カテゴリ振り分け ===

def determine_sheet(item):
    """変換済みアイテムをどのシートに振り分けるか判定"""
    katashiki = item.get('型式', '')
    text_color = item.get('文字カラー', '')
    qty_type = item.get('複数', '単品')
    font = item.get('書体', '')
    name = item.get('作成名変換', '')
    memo = item.get('ひとことメモ', '')

    # クルトガ系
    if 'クルトガ' in katashiki:
        if qty_type == '単品':
            return 'クルトガ単品'
        elif qty_type == '単品+':
            return 'クルトガ単品+'
        else:
            return 'クルトガ複数'

    # ピュアブラック 4&1
    if 'ピュアブラック' in katashiki:
        return 'ピュアブラック 4&1'

    # ピュアモルトシングル
    if 'ピュアシングル' in katashiki:
        if qty_type in ('単品', '単品+'):
            return 'ピュアモルトシングル単品'
        else:
            return 'ピュアモルトシングル複数'

    # ラミー/その他特殊型式
    if 'ラミー' in katashiki or 'プライム' in katashiki:
        if font == '名入れ不要' or not name:
            return 'その他2名入れなし'
        return 'その他3名入れあり'

    # 欠品チェック
    if '欠品' in memo:
        return 'その他1 欠品'

    # 通常ジェットストリーム
    if qty_type == '単品':
        if text_color == 'ブラック':
            return '単品 ブラック'
        elif text_color == 'ホワイト':
            return '単品 ホワイト'
        else:
            return 'その他3名入れあり' if name else 'その他2名入れなし'
    elif qty_type == '単品+':
        return '単品+'
    else:
        return '複数'


# === Excel出力 ===

def write_excel(items, output_path):
    """変換済みデータをExcelに出力"""
    wb = openpyxl.Workbook()
    header_font = Font(bold=True)

    # Sheet1（全データ）
    ws1 = wb.active
    ws1.title = 'Sheet1'
    for c, h in enumerate(HEADERS, 1):
        ws1.cell(1, c, h).font = header_font

    for r, item in enumerate(items, 2):
        for c, h in enumerate(HEADERS, 1):
            val = item.get(h, '')
            if h == '文字数' and val:
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    pass
            ws1.cell(r, c, val)

    print(f"  Sheet1: {len(items)}行")

    # カテゴリ別シート
    sheet_names = [
        '単品 ブラック', '単品 ホワイト', '単品+', '複数',
        'ピュアブラック 4&1',
        'クルトガ単品', 'クルトガ単品+', 'クルトガ複数',
        'ピュアモルトシングル単品', 'ピュアモルトシングル複数',
        'その他3名入れあり', 'その他1 欠品', 'その他2名入れなし',
    ]

    by_sheet = defaultdict(list)
    for item in items:
        sheet = determine_sheet(item)
        by_sheet[sheet].append(item)

    for sn in sheet_names:
        ws = wb.create_sheet(sn)
        for c, h in enumerate(HEADERS, 1):
            ws.cell(1, c, h).font = header_font

        sheet_items = by_sheet.get(sn, [])
        for r, item in enumerate(sheet_items, 2):
            for c, h in enumerate(HEADERS, 1):
                val = item.get(h, '')
                if h == '文字数' and val:
                    try:
                        val = int(float(val))
                    except (ValueError, TypeError):
                        pass
                ws.cell(r, c, val)

        if sheet_items:
            print(f"  {sn}: {len(sheet_items)}行")

    wb.save(output_path)
    print(f"\n出力: {output_path}")


# === メイン ===

def main():
    parser = argparse.ArgumentParser(description='筆記具割付表自動変換')
    parser.add_argument('rakuten_csv', help='楽天CSV')
    parser.add_argument('yahoo_csv', help='Yahoo CSV')
    parser.add_argument('amazon_csv', help='Amazon CSV')
    parser.add_argument('--master', required=True, help='③割付表ver5ファイル (.xls)')
    parser.add_argument('--db', required=True, help='①振り分け表ファイル (.xls)')
    parser.add_argument('--output', default=None, help='出力ファイル名')
    args = parser.parse_args()

    print("マスターデータ読み込み...")
    master = load_master_data(args.master)
    product_db = load_product_db(args.db)
    print(f"  文字カラー: {len(master['color_map'])}件")
    print(f"  テンプレート: {len(master['template_map'])}件")
    print(f"  型式: {len(master['type_map'])}件")
    print(f"  イラスト変換: {len(master['illust_map'])}件")
    print(f"  商品DB: {len(product_db)}件")

    print("\nCSV読み込み...")
    all_rows = []
    for path, channel in [
        (args.rakuten_csv, '楽天'),
        (args.yahoo_csv, 'Yahoo'),
        (args.amazon_csv, 'Amazon'),
    ]:
        if not os.path.exists(path):
            print(f"  警告: {path} が見つかりません。スキップします。")
            continue
        rows = read_csv_rows(path, channel)
        rows = normalize_columns(rows, channel)
        all_rows.extend(rows)

    # 筆記具フィルタ
    pen_rows = [r for r in all_rows if is_pen_item(r, product_db)]
    print(f"\n筆記具フィルタ: {len(all_rows)}→{len(pen_rows)}行")

    # 変換
    print("\n変換処理...")
    items = []
    for row in pen_rows:
        item = convert_row(row, master, product_db)
        items.append(item)

    # 出力
    output = args.output or '割付表_筆記具_自動生成.xlsx'
    print(f"\nExcel出力...")
    write_excel(items, output)


if __name__ == '__main__':
    main()
