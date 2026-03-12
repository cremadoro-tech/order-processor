"""
generate_proposal.py
ハンコヤストア 企画書自動生成スクリプト

使い方:
    python generate_proposal.py input.json [--strategy path/to/ec-strategy.md] [--lp path/to/lp.html]

input.json は Claude Code が出力する JSON ファイル。
sample_input.json を参照してください。
"""

import json
import sys
import os
import re
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "hankoya_proposal_template.xlsx")

def fill_proposal(data: dict, output_path: str):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["企画書"]

    def write(row, col, value):
        """指定セルに値を書き込む（結合セルは先頭セルに書く）"""
        ws.cell(row=row, column=col, value=value)

    # ── タイトル ──
    title = f"【 {data.get('item_code', '')}：{data.get('item_name', '')} 】"
    ws.cell(row=1, column=2, value=title)

    # ── 01 商品概要 ──
    write(3, 3, data.get("monthly_target", ""))
    write(3, 7, data.get("competitor_url", ""))
    write(4, 3, data.get("nint_monthly_avg", ""))
    write(4, 7, data.get("nint_peak_month", ""))
    write(5, 3, data.get("adoption_reason", ""))

    # ── 03 ターゲット・ペルソナ ──
    persona = data.get("persona", {})
    write(9,  3, persona.get("buyer", ""))
    write(10, 3, persona.get("recipient", ""))
    write(11, 3, persona.get("target_event", ""))
    write(12, 3, persona.get("reason_to_choose", ""))
    write(13, 3, persona.get("name_engraving_example", ""))

    # ── 04 競合・市場分析 ──
    market = data.get("market_analysis", {})
    write(15, 3, market.get("market_size_trend", ""))
    write(16, 3, market.get("main_competitors", ""))
    write(17, 3, market.get("differentiation", ""))
    write(18, 3, market.get("product_strengths", ""))

    # ── 08 発注情報 ──
    orders = data.get("order_info", [])
    order_start_row = None
    # 発注情報開始行を動的に探す
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "NE品番":
                order_start_row = cell.row + 1
                break
        if order_start_row:
            break

    if order_start_row:
        for i, order in enumerate(orders[:3]):
            r = order_start_row + i
            ws.cell(row=r, column=2, value=order.get("ne_code", ""))
            ws.cell(row=r, column=3, value=order.get("jan_code", ""))
            ws.cell(row=r, column=4, value=order.get("maker_model", ""))
            ws.cell(row=r, column=5, value=order.get("item_name", ""))
            ws.cell(row=r, column=6, value=order.get("lot", ""))
            ws.cell(row=r, column=7, value=order.get("initial_order_qty", ""))
            ws.cell(row=r, column=8, value=order.get("sample_qty", "無し"))
            ws.cell(row=r, column=9, value=order.get("wholesale_price", ""))

    # wb を返す（main で追加タブ処理後にまとめて保存）
    return wb


def _save_workbook(wb, output_path: str):
    wb.save(output_path)
    print(f"✅ 企画書を生成しました: {output_path}")


def add_strategy_tab(wb: openpyxl.Workbook, md_path: str):
    """EC戦略MDファイルをExcelの新規タブに書き込む"""
    tab_name = "EC戦略"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)

    h1_font     = Font(bold=True, size=14)
    h2_font     = Font(bold=True, size=12)
    h3_font     = Font(bold=True, size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    sub_fill    = PatternFill(fill_type="solid", fgColor="EBF0FA")
    wrap_align  = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45

    with open(md_path, encoding="utf-8") as f:
        lines = f.readlines()

    row = 1
    for line in lines:
        line = line.rstrip("\n")

        if line.startswith("### "):
            c = ws.cell(row=row, column=1, value=line[4:])
            c.font = h3_font
            c.fill = sub_fill
            c.alignment = wrap_align
            row += 1
        elif line.startswith("## "):
            c = ws.cell(row=row, column=1, value=line[3:])
            c.font = h2_font
            c.fill = header_fill
            c.alignment = wrap_align
            row += 1
        elif line.startswith("# "):
            c = ws.cell(row=row, column=1, value=line[2:])
            c.font = h1_font
            c.alignment = wrap_align
            row += 1
        elif line.startswith("|"):
            cells = [c.strip() for c in line.split("|")[1:-1]]
            # 区切り行はスキップ
            if all(re.match(r"^[-: ]+$", c) for c in cells if c):
                continue
            for col_idx, val in enumerate(cells[:4], start=1):
                plain = re.sub(r"\*\*(.+?)\*\*", r"\1", val)
                cell = ws.cell(row=row, column=col_idx, value=plain)
                cell.alignment = wrap_align
            row += 1
        elif line.strip() == "":
            row += 1
        else:
            plain = re.sub(r"\*\*(.+?)\*\*", r"\1", line)
            plain = re.sub(r"`(.+?)`", r"\1", plain)
            c = ws.cell(row=row, column=1, value=plain)
            c.alignment = wrap_align
            row += 1

    print(f"✅ EC戦略タブを追加しました: {tab_name}")


def add_lp_tab(wb: openpyxl.Workbook, lp_path: str):
    """LP HTMLファイルをExcelの新規タブに書き込む（テキスト抽出）"""
    tab_name = "LP"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)

    ws.column_dimensions["A"].width = 120
    wrap_align = Alignment(wrap_text=True, vertical="top")

    with open(lp_path, encoding="utf-8") as f:
        content = f.read()

    # HTMLタグを除去してテキスト抽出
    text = re.sub(r"<style[^>]*>.*?</style>", "", content, flags=re.DOTALL)
    text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    ws.cell(row=1, column=1, value="LP本文（HTMLより抽出）").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"ソース: {os.path.basename(lp_path)}")

    for i, line in enumerate(text.splitlines(), start=3):
        line = line.strip()
        if line:
            ws.cell(row=i, column=1, value=line).alignment = wrap_align
            ws.row_dimensions[i].height = 18

    print(f"✅ LPタブを追加しました: {tab_name}")


def main():
    if len(sys.argv) < 2:
        print("使い方: python generate_proposal.py input.json [--strategy path.md] [--lp path.html]")
        sys.exit(1)

    json_path = sys.argv[1]
    args = sys.argv[2:]

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    item_code = data.get("item_code", "XXXX").replace("/", "-")
    output_path = f"企画書_{item_code}_{timestamp}.xlsx"

    strategy_path = None
    lp_path = None
    for i, arg in enumerate(args):
        if arg == "--strategy" and i + 1 < len(args):
            strategy_path = args[i + 1]
        if arg == "--lp" and i + 1 < len(args):
            lp_path = args[i + 1]

    # テンプレートロード→タブ追加→1回だけ保存
    wb = fill_proposal(data, output_path)
    if strategy_path and os.path.exists(strategy_path):
        add_strategy_tab(wb, strategy_path)
    if lp_path and os.path.exists(lp_path):
        add_lp_tab(wb, lp_path)
    _save_workbook(wb, output_path)

    print(f"📄 保存先: {output_path}")


if __name__ == "__main__":
    main()
